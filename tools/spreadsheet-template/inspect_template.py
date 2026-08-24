from openpyxl import load_workbook


def main():
    path = r"C:\Users\aparna_adevtechcorp\Downloads\Aparana- task tracker (1).xlsx"
    wb = load_workbook(path)
    print("sheets", wb.sheetnames)
    for ws in wb.worksheets:
        print(ws.title, ws.max_row, ws.max_column, "merged", [str(r) for r in ws.merged_cells.ranges])
        print("widths", [(k, v.width) for k, v in ws.column_dimensions.items() if v.width][:20])
        print("heights", [(k, v.height) for k, v in ws.row_dimensions.items() if v.height][:20])
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), min_col=1, max_col=min(ws.max_column, 12)):
            print([c.value for c in row])
        print("---")


if __name__ == "__main__":
    main()
