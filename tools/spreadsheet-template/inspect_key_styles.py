import json
from openpyxl import load_workbook

from build_sheet_requests import SOURCE, cell_format


def main():
    wb = load_workbook(SOURCE, data_only=False)
    for ws in wb.worksheets:
        print(ws.title)
        for addr in ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "A2", "F2", "G2", "E4", "D12"]:
            print(addr, json.dumps(cell_format(ws[addr]), sort_keys=True))


if __name__ == "__main__":
    main()
