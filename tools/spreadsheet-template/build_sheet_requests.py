import json
from copy import copy
from datetime import datetime, time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SOURCE = r"C:\Users\aparna_adevtechcorp\Downloads\Spreadsheets\Aparana- task tracker (1).xlsx"
TARGET_SHEETS = {"August": 0, "Sept": 1234567}


def rgb_color(color):
    if not color or color.type != "rgb" or not color.rgb:
        return None
    value = color.rgb[-6:]
    return {
        "red": int(value[0:2], 16) / 255,
        "green": int(value[2:4], 16) / 255,
        "blue": int(value[4:6], 16) / 255,
    }


def text_format(font):
    fmt = {}
    if font.name:
        fmt["fontFamily"] = font.name
    if font.sz:
        fmt["fontSize"] = int(font.sz)
    if font.bold is not None:
        fmt["bold"] = bool(font.bold)
    if font.italic is not None:
        fmt["italic"] = bool(font.italic)
    if font.underline:
        fmt["underline"] = True
    color = rgb_color(font.color)
    if color:
        fmt["foregroundColor"] = color
    return fmt


def border_side(side):
    if not side or not side.style:
        return None
    style_map = {
        "thin": "SOLID",
        "medium": "SOLID_MEDIUM",
        "thick": "SOLID_THICK",
        "dashed": "DASHED",
        "dotted": "DOTTED",
        "double": "DOUBLE",
        "hair": "DOTTED",
    }
    out = {"style": style_map.get(side.style, "SOLID")}
    color = rgb_color(side.color)
    if color:
        out["color"] = color
    return out


def cell_format(cell):
    fmt = {}
    fill = cell.fill
    if fill and fill.fill_type == "solid":
        color = rgb_color(fill.fgColor)
        if color:
            fmt["backgroundColor"] = color
    tf = text_format(cell.font)
    if tf:
        fmt["textFormat"] = tf
    if cell.alignment:
        if cell.alignment.horizontal:
            fmt["horizontalAlignment"] = cell.alignment.horizontal.upper().replace("CENTER", "CENTER")
        if cell.alignment.vertical:
            fmt["verticalAlignment"] = cell.alignment.vertical.upper()
        if cell.alignment.wrap_text:
            fmt["wrapStrategy"] = "WRAP"
    if cell.number_format and cell.number_format != "General":
        pattern = cell.number_format
        fmt["numberFormat"] = {
            "type": "DATE_TIME" if "h" in pattern.lower() and "d" in pattern.lower() else "NUMBER",
            "pattern": pattern,
        }
        if "d" in pattern.lower() and "y" in pattern.lower():
            fmt["numberFormat"]["type"] = "DATE"
        elif "h" in pattern.lower() or "s" in pattern.lower():
            fmt["numberFormat"]["type"] = "TIME"
    borders = {}
    for name in ("top", "bottom", "left", "right"):
        side = border_side(getattr(cell.border, name))
        if side:
            borders[name] = side
    if borders:
        fmt["borders"] = borders
    return fmt


def entered_value(cell):
    value = cell.value
    if value is None:
        return None
    if cell.data_type == "f":
        return {"formulaValue": f"={value}" if not str(value).startswith("=") else str(value)}
    if isinstance(value, datetime):
        # Google Sheets interprets DATE number formats against spreadsheet serials.
        serial = (value - datetime(1899, 12, 30)).days
        return {"numberValue": serial}
    if isinstance(value, time):
        seconds = value.hour * 3600 + value.minute * 60 + value.second
        return {"numberValue": seconds / 86400}
    if isinstance(value, bool):
        return {"boolValue": value}
    if isinstance(value, (int, float)):
        return {"numberValue": value}
    return {"stringValue": str(value)}


def dimension_requests(ws, sheet_id):
    requests = []
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        width = ws.column_dimensions[letter].width
        if width:
            requests.append({
                "updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": idx - 1, "endIndex": idx},
                    "properties": {"pixelSize": max(20, int(width * 7))},
                    "fields": "pixelSize",
                }
            })
    for idx in range(1, ws.max_row + 1):
        height = ws.row_dimensions[idx].height
        if height:
            requests.append({
                "updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": idx - 1, "endIndex": idx},
                    "properties": {"pixelSize": max(15, int(height * 1.33))},
                    "fields": "pixelSize",
                }
            })
    return requests


def sheet_requests(ws, sheet_id):
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        values = []
        for cell in row:
            data = {}
            value = entered_value(cell)
            if value:
                data["userEnteredValue"] = value
            fmt = cell_format(cell)
            if fmt:
                data["userEnteredFormat"] = fmt
            values.append(data)
        rows.append({"values": values})
    requests = [{
        "updateCells": {
            "start": {"sheetId": sheet_id, "rowIndex": 0, "columnIndex": 0},
            "rows": rows,
            "fields": "userEnteredValue,userEnteredFormat",
        }
    }]
    for merged in ws.merged_cells.ranges:
        requests.append({
            "mergeCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": merged.min_row - 1,
                    "endRowIndex": merged.max_row,
                    "startColumnIndex": merged.min_col - 1,
                    "endColumnIndex": merged.max_col,
                },
                "mergeType": "MERGE_ALL",
            }
        })
    requests.extend(dimension_requests(ws, sheet_id))
    return requests


def main():
    wb = load_workbook(SOURCE, data_only=False)
    requests = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": 0,
                    "title": "August",
                    "gridProperties": {"rowCount": wb["August"].max_row, "columnCount": wb["August"].max_column},
                },
                "fields": "title,gridProperties.rowCount,gridProperties.columnCount",
            }
        },
        {
            "addSheet": {
                "properties": {
                    "sheetId": TARGET_SHEETS["Sept"],
                    "title": "Sept",
                    "gridProperties": {"rowCount": wb["Sept"].max_row, "columnCount": wb["Sept"].max_column},
                }
            }
        },
    ]
    requests.extend(sheet_requests(wb["August"], TARGET_SHEETS["August"]))
    requests.extend(sheet_requests(wb["Sept"], TARGET_SHEETS["Sept"]))
    print(json.dumps(requests))


if __name__ == "__main__":
    main()
