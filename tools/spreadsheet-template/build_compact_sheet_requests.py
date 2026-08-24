import json
from copy import deepcopy
from datetime import datetime, time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from build_sheet_requests import SOURCE, TARGET_SHEETS, cell_format


def tsv_value(cell):
    value = cell.value
    if value is None:
        return ""
    if cell.data_type == "f":
        text = str(value)
        return text if text.startswith("=") else f"={text}"
    if isinstance(value, datetime):
        return value.strftime("%d-%b-%y")
    if isinstance(value, time):
        hour = value.hour
        suffix = "AM" if hour < 12 else "PM"
        hour12 = hour % 12 or 12
        return f"{hour12}:{value.minute:02d} {suffix}"
    return str(value)


def clean_format(fmt):
    fmt = deepcopy(fmt)
    text = fmt.get("textFormat")
    if text:
        if text.get("fontFamily") == "Calibri":
            text.pop("fontFamily", None)
        if text.get("fontSize") == 11:
            text.pop("fontSize", None)
        if text.get("bold") is False:
            text.pop("bold", None)
        if text.get("italic") is False:
            text.pop("italic", None)
        if not text:
            fmt.pop("textFormat", None)
    return fmt


def row_style_runs(ws, sheet_id):
    requests = []
    for row_idx in range(1, ws.max_row + 1):
        run_fmt = None
        run_start = 1
        for col_idx in range(1, ws.max_column + 2):
            fmt = None
            if col_idx <= ws.max_column:
                fmt = clean_format(cell_format(ws.cell(row_idx, col_idx)))
            if fmt != run_fmt:
                if run_fmt:
                    requests.append({
                        "repeatCell": {
                            "range": {
                                "sheetId": sheet_id,
                                "startRowIndex": row_idx - 1,
                                "endRowIndex": row_idx,
                                "startColumnIndex": run_start - 1,
                                "endColumnIndex": col_idx - 1,
                            },
                            "cell": {"userEnteredFormat": run_fmt},
                            "fields": "userEnteredFormat",
                        }
                    })
                run_fmt = fmt
                run_start = col_idx
    return requests


def dimension_requests(ws, sheet_id):
    requests = []
    for idx in range(1, ws.max_column + 1):
        width = ws.column_dimensions[get_column_letter(idx)].width
        if width:
            requests.append({
                "updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": idx - 1, "endIndex": idx},
                    "properties": {"pixelSize": max(20, int(width * 7))},
                    "fields": "pixelSize",
                }
            })
    for idx in range(1, ws.max_row + 1):
        height = ws.row_dimensions[idx].height
        if height:
            requests.append({
                "updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": idx - 1, "endIndex": idx},
                    "properties": {"pixelSize": max(15, int(height * 1.33))},
                    "fields": "pixelSize",
                }
            })
    return requests


def sheet_requests(ws, sheet_id):
    data = "\n".join("\t".join(tsv_value(cell) for cell in row) for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ))
    requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": ws.max_row,
                    "startColumnIndex": 0,
                    "endColumnIndex": ws.max_column,
                },
                "cell": {},
                "fields": "userEnteredValue,userEnteredFormat",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": ws.max_row,
                    "startColumnIndex": 0,
                    "endColumnIndex": ws.max_column,
                },
                "cell": {"userEnteredFormat": {"textFormat": {"fontFamily": "Calibri", "fontSize": 11}}},
                "fields": "userEnteredFormat.textFormat",
            }
        },
        {
            "pasteData": {
                "coordinate": {"sheetId": sheet_id, "rowIndex": 0, "columnIndex": 0},
                "data": data,
                "type": "PASTE_NORMAL",
                "delimiter": "\t",
            }
        },
    ]
    requests.extend(row_style_runs(ws, sheet_id))
    for merged in ws.merged_cells.ranges:
        requests.append({
            "mergeCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": merged.min_row - 1,
                    "endRowIndex": merged.max_row,
                    "startColumnIndex": merged.min_col - 1,
                    "endColumnIndex": merged.max_col,
                },
                "mergeType": "MERGE_ALL",
            }
        })
    requests.extend(dimension_requests(ws, sheet_id))
    return requests


def main():
    wb = load_workbook(SOURCE, data_only=False)
    requests = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": 0,
                    "title": "August",
                    "gridProperties": {"rowCount": wb["August"].max_row, "columnCount": wb["August"].max_column},
                },
                "fields": "title,gridProperties.rowCount,gridProperties.columnCount",
            }
        },
        {
            "addSheet": {
                "properties": {
                    "sheetId": TARGET_SHEETS["Sept"],
                    "title": "Sept",
                    "gridProperties": {"rowCount": wb["Sept"].max_row, "columnCount": wb["Sept"].max_column},
                }
            }
        },
    ]
    requests.extend(sheet_requests(wb["August"], TARGET_SHEETS["August"]))
    requests.extend(sheet_requests(wb["Sept"], TARGET_SHEETS["Sept"]))
    print(json.dumps(requests, separators=(",", ":")))


if __name__ == "__main__":
    main()
