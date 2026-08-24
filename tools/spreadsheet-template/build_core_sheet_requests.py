import json
from datetime import datetime, time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from build_sheet_requests import SOURCE, TARGET_SHEETS


MEDIUM = {"style": "SOLID_MEDIUM"}
THIN = {"style": "SOLID"}


def tsv_value(cell):
    value = cell.value
    if value is None:
        return ""
    if cell.data_type == "f":
        text = str(value)
        return text if text.startswith("=") else f"={text}"
    if isinstance(value, datetime):
        return value.strftime("%d-%b-%y")
    if isinstance(value, time):
        hour = value.hour % 12 or 12
        suffix = "AM" if value.hour < 12 else "PM"
        return f"{hour}:{value.minute:02d} {suffix}"
    return str(value).replace("\t", " ").replace("\n", " ")


def repeat(sheet_id, start_row, end_row, start_col, end_col, fmt, fields="userEnteredFormat"):
    return {
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row,
                "endRowIndex": end_row,
                "startColumnIndex": start_col,
                "endColumnIndex": end_col,
            },
            "cell": {"userEnteredFormat": fmt},
            "fields": fields,
        }
    }


def sheet_requests(ws, sheet_id):
    data = "\n".join(
        "\t".join(tsv_value(cell) for cell in row)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
    )
    requests = [
        repeat(sheet_id, 0, ws.max_row, 0, ws.max_column, {}, "userEnteredValue,userEnteredFormat"),
        repeat(sheet_id, 0, ws.max_row, 0, ws.max_column, {
            "textFormat": {"fontFamily": "Calibri", "fontSize": 11},
            "verticalAlignment": "MIDDLE",
        }, "userEnteredFormat.textFormat,userEnteredFormat.verticalAlignment"),
        {
            "pasteData": {
                "coordinate": {"sheetId": sheet_id, "rowIndex": 0, "columnIndex": 0},
                "data": data,
                "type": "PASTE_NORMAL",
                "delimiter": "\t",
            }
        },
        repeat(sheet_id, 0, 1, 0, ws.max_column, {
            "textFormat": {"fontFamily": "Calibri", "fontSize": 12, "bold": True},
            "horizontalAlignment": "CENTER",
            "verticalAlignment": "MIDDLE",
            "wrapStrategy": "WRAP",
        }, "userEnteredFormat.textFormat,userEnteredFormat.horizontalAlignment,userEnteredFormat.verticalAlignment,userEnteredFormat.wrapStrategy"),
        repeat(sheet_id, 1, ws.max_row, 0, 2, {"horizontalAlignment": "CENTER"}, "userEnteredFormat.horizontalAlignment"),
        repeat(sheet_id, 1, ws.max_row, 4, 7, {"horizontalAlignment": "CENTER"}, "userEnteredFormat.horizontalAlignment"),
        repeat(sheet_id, 1, ws.max_row, 0, 1, {"numberFormat": {"type": "DATE", "pattern": "d-mmm-yy"}}, "userEnteredFormat.numberFormat"),
        repeat(sheet_id, 1, ws.max_row, 4, 6, {"numberFormat": {"type": "TIME", "pattern": "h:mm"}}, "userEnteredFormat.numberFormat"),
        repeat(sheet_id, 1, ws.max_row, 3, 4, {"wrapStrategy": "WRAP"}, "userEnteredFormat.wrapStrategy"),
        {
            "updateBorders": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": ws.max_row, "startColumnIndex": 0, "endColumnIndex": ws.max_column},
                "top": MEDIUM,
                "bottom": MEDIUM,
                "left": MEDIUM,
                "right": MEDIUM,
                "innerHorizontal": THIN,
                "innerVertical": MEDIUM,
            }
        },
    ]
    for merged in ws.merged_cells.ranges:
        requests.append({
            "mergeCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": merged.min_row - 1,
                    "endRowIndex": merged.max_row,
                    "startColumnIndex": merged.min_col - 1,
                    "endColumnIndex": merged.max_col,
                },
                "mergeType": "MERGE_ALL",
            }
        })
    for idx in range(1, ws.max_column + 1):
        width = ws.column_dimensions[get_column_letter(idx)].width
        if width:
            requests.append({
                "updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": idx - 1, "endIndex": idx},
                    "properties": {"pixelSize": max(20, int(width * 7))},
                    "fields": "pixelSize",
                }
            })
    for idx in range(1, ws.max_row + 1):
        height = ws.row_dimensions[idx].height
        if height:
            requests.append({
                "updateDimensionProperties": {
                    "range": {"sheetId": sheet_id, "dimension": "ROWS", "startIndex": idx - 1, "endIndex": idx},
                    "properties": {"pixelSize": max(15, int(height * 1.33))},
                    "fields": "pixelSize",
                }
            })
    return requests


def main():
    wb = load_workbook(SOURCE, data_only=False)
    requests = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": 0,
                    "title": "August",
                    "gridProperties": {"rowCount": wb["August"].max_row, "columnCount": wb["August"].max_column},
                },
                "fields": "title,gridProperties.rowCount,gridProperties.columnCount",
            }
        },
        {
            "addSheet": {
                "properties": {
                    "sheetId": TARGET_SHEETS["Sept"],
                    "title": "Sept",
                    "gridProperties": {"rowCount": wb["Sept"].max_row, "columnCount": wb["Sept"].max_column},
                }
            }
        },
    ]
    requests.extend(sheet_requests(wb["August"], TARGET_SHEETS["August"]))
    requests.extend(sheet_requests(wb["Sept"], TARGET_SHEETS["Sept"]))
    print(json.dumps(requests, separators=(",", ":")))


if __name__ == "__main__":
    main()
