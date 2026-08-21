import json
from collections import Counter

from openpyxl import load_workbook

from build_sheet_requests import SOURCE, cell_format


def main():
    wb = load_workbook(SOURCE, data_only=False)
    for ws in wb.worksheets:
        counts = Counter()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                counts[json.dumps(cell_format(cell), sort_keys=True)] += 1
        print(ws.title, "unique styles", len(counts))
        for style, count in counts.most_common(10):
            print(count, style[:250])


if __name__ == "__main__":
    main()
