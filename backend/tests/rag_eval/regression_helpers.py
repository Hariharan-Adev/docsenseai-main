"""Reusable local fixtures and assertions for known RAG failure patterns."""

from __future__ import annotations

import asyncio
from contextlib import ExitStack
from dataclasses import dataclass
from io import BytesIO
import json
from pathlib import Path
import re
import tempfile
from typing import Any
import unittest
from unittest.mock import patch

from fastapi import UploadFile
from openpyxl import Workbook
from starlette.requests import Request

from app import database
from app.prompts.rag_prompt import UNAVAILABLE_ANSWER
from app.routes import upload
from app.services import rag_service, source_selection, structured_ingestion, vector_store
from app.services.chat_context import resolve_follow_up
from app.services.source_selection import select_sources
from app.services.vector_store import reset_vector_store_for_tests
from app.services.workbook_analysis import (
    analyze_workbook_question,
    has_structured_workbook,
    is_analytical_question,
    is_structured_lookup_question,
)
from tests.rag_eval.helpers import RetrievalObservation, RetrievedChunk


REGRESSION_CASE_FILE = Path(__file__).with_name("regression_cases.json")
AGRICULTURE_FIXTURE = Path(__file__).parents[1] / "fixtures" / "agriculture_dataset.csv"
REQUIRED_CASE_IDS = {
    "multi_tab_inventory_lookup",
    "exact_person_name_lookup",
    "generic_record_count",
    "generic_percentage_aggregation",
    "generic_numeric_range",
    "generic_exact_value_lookup",
    "named_workbook_source_scope",
    "missing_fact_is_unavailable",
    "follow_up_name_them",
    "follow_up_other_ones",
    "follow_up_show_those",
    "follow_up_without_context_stays_local",
    "multiple_workbook_schema_selection",
    "ambiguous_workbook_requires_selection",
}


@dataclass(frozen=True)
class RetrievalExpectation:
    """Expected behavior before final answer assertions are applied."""

    route: str
    documents: tuple[str, ...]
    sheets: tuple[str, ...] = ()
    forbid_global_search: bool = False


@dataclass(frozen=True)
class AnswerExpectation:
    """Expected grounded answer behavior independent of retrieval ranking."""

    grounded: bool
    unavailable: bool
    facts: tuple[str, ...]
    excluded_facts: tuple[str, ...] = ()
    question_type: str | None = None


@dataclass(frozen=True)
class PriorTurn:
    """One query used only to establish bounded conversation context."""

    query: str
    expected_facts: tuple[str, ...] = ()
    expected_documents: tuple[str, ...] = ()


@dataclass(frozen=True)
class RagRegressionCase:
    """One portable regression with separate retrieval and answer contracts."""

    id: str
    pattern: str
    fixture: str
    query: str
    expected_retrieval: RetrievalExpectation
    expected_answer: AnswerExpectation
    prior_turns: tuple[PriorTurn, ...] = ()

    @classmethod
    def from_dict(cls, value: dict[str, Any]) -> "RagRegressionCase":
        """Parse and validate a nested regression case from JSON."""
        retrieval = value["expected_retrieval"]
        answer = value["expected_answer"]
        case = cls(
            id=str(value["id"]),
            pattern=str(value["pattern"]),
            fixture=str(value["fixture"]),
            query=str(value["query"]),
            expected_retrieval=RetrievalExpectation(
                route=str(retrieval["route"]),
                documents=tuple(str(item) for item in retrieval.get("documents", [])),
                sheets=tuple(str(item) for item in retrieval.get("sheets", [])),
                forbid_global_search=bool(retrieval.get("forbid_global_search", False)),
            ),
            expected_answer=AnswerExpectation(
                grounded=bool(answer["grounded"]),
                unavailable=bool(answer["unavailable"]),
                facts=tuple(str(item) for item in answer.get("facts", [])),
                excluded_facts=tuple(str(item) for item in answer.get("excluded_facts", [])),
                question_type=(
                    str(answer["question_type"])
                    if answer.get("question_type") is not None
                    else None
                ),
            ),
            prior_turns=tuple(
                PriorTurn(
                    query=str(item["query"]),
                    expected_facts=tuple(str(fact) for fact in item.get("expected_facts", [])),
                    expected_documents=tuple(
                        str(filename) for filename in item.get("expected_documents", [])
                    ),
                )
                for item in value.get("prior_turns", [])
            ),
        )
        case.validate()
        return case

    def validate(self) -> None:
        """Reject incomplete or contradictory baseline expectations early."""
        if not all((self.id.strip(), self.pattern.strip(), self.fixture.strip(), self.query.strip())):
            raise ValueError("RAG regression identity, pattern, fixture, and query are required.")
        if self.expected_answer.unavailable and self.expected_answer.facts:
            raise ValueError(f"Unavailable regression {self.id} cannot require answer facts.")
        if self.expected_retrieval.forbid_global_search and self.prior_turns:
            raise ValueError(f"Regression {self.id} cannot forbid search after seeded prior turns.")


@dataclass(frozen=True)
class RetrievalRun:
    """Observed retrieval result plus whether fallback global search ran."""

    observation: RetrievalObservation
    sheets: tuple[str, ...]
    global_search_attempted: bool = False


def load_regression_cases(path: Path = REGRESSION_CASE_FILE) -> list[RagRegressionCase]:
    """Load local regression cases without sending fixture data externally."""
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        raise ValueError("RAG regression cases must be a JSON array.")
    cases = [RagRegressionCase.from_dict(item) for item in raw]
    if len({case.id for case in cases}) != len(cases):
        raise ValueError("RAG regression case IDs must be unique.")
    missing = REQUIRED_CASE_IDS - {case.id for case in cases}
    if missing:
        raise ValueError(f"Required RAG regression cases are missing: {sorted(missing)}")
    return cases


def _request() -> Request:
    """Build the minimal request object required by the upload test helper."""
    return Request({
        "type": "http",
        "method": "POST",
        "path": "/documents/upload",
        "headers": [],
        "client": ("rag-eval", 1),
    })


def _workbook_bytes(sheets: list[tuple[str, list[list[object]]]]) -> bytes:
    """Create a small workbook fixture in memory to avoid production data."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets:
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class RagRegressionHarness:
    """Own an isolated SQLite/upload environment for one regression test."""

    def __init__(self) -> None:
        """Initialize fields before the test opens temporary resources."""
        self.temporary = tempfile.TemporaryDirectory()
        self.stack = ExitStack()
        self.uploaded: dict[str, int] = {}
        self.conversation_id = "rag-eval-conversation"

    def start(self) -> None:
        """Patch only test storage and deterministic embedding boundaries."""
        root = Path(self.temporary.name)
        database_path = root / "rag-eval.db"
        upload_path = root / "uploads"
        self.stack.enter_context(patch.object(database, "DATABASE_PATH", database_path))
        self.stack.enter_context(patch.object(database, "UPLOAD_DIRECTORY", upload_path))
        self.stack.enter_context(patch.object(vector_store.settings, "vector_store", "sqlite"))
        self.stack.enter_context(patch.object(vector_store.settings, "vector_store_provider", "sqlite"))
        self.stack.enter_context(patch.object(vector_store.settings, "qdrant_local_path", ""))
        self.stack.enter_context(patch.object(upload, "UPLOAD_DIRECTORY", upload_path))
        self.stack.enter_context(patch.object(upload, "enforce_request_limit", lambda *args, **kwargs: None))
        self.stack.enter_context(patch.object(upload, "log_audit_event", lambda **kwargs: None))
        self.stack.enter_context(patch.object(source_selection, "log_event", lambda *args, **kwargs: None))
        self.stack.enter_context(
            patch.object(
                upload,
                "create_embeddings",
                lambda chunks: [[1.0, 0.0] + [0.0] * 382 for _ in chunks],
            )
        )
        database.initialize_database()
        reset_vector_store_for_tests()
        with database.get_connection() as connection:
            connection.execute(
                "INSERT INTO users (id, email, password_hash) VALUES (1, 'rag-eval@example.com', 'hash')"
            )

    def close(self) -> None:
        """Release patches before deleting the isolated temporary directory."""
        self.stack.close()
        self.temporary.cleanup()

    def upload_workbook(self, filename: str, sheets: list[tuple[str, list[list[object]]]]) -> int:
        """Upload an in-memory workbook through the existing ingestion path."""
        file = UploadFile(file=BytesIO(_workbook_bytes(sheets)), filename=filename)
        result = asyncio.run(upload._process_document_upload(_request(), file, {"id": 1}))
        document_id = int(result["document_id"])
        self.uploaded[filename] = document_id
        return document_id

    def upload_csv_fixture(self) -> int:
        """Reuse and fully structured-index the repository agriculture CSV fixture."""
        file = UploadFile(
            file=BytesIO(AGRICULTURE_FIXTURE.read_bytes()),
            filename=AGRICULTURE_FIXTURE.name,
        )
        result = asyncio.run(upload._process_document_upload(_request(), file, {"id": 1}))
        document_id = int(result["document_id"])
        with database.get_connection() as connection:
            organization_id = str(
                connection.execute("SELECT organization_id FROM users WHERE id = 1").fetchone()[0]
            )
        # The synchronous upload helper predates CSV structured indexing; mirror the
        # current ingestion worker step so this fixture tests retrieval, not setup.
        structured_ingestion.reindex_existing_csv_document(
            document_id=document_id,
            owner_id=1,
            organization_id=organization_id,
        )
        self.uploaded[AGRICULTURE_FIXTURE.name] = document_id
        return document_id

    def prepare(self, fixture: str) -> None:
        """Load one named, domain-neutral fixture scenario."""
        builders = {
            "multi_tab_inventory": self._prepare_multi_tab_inventory,
            "people_profiles": self._prepare_people_profiles,
            "project_roster": self._prepare_project_roster,
            "campaign_metrics": self._prepare_campaign_metrics,
            "agriculture_csv": self.upload_csv_fixture,
            "source_specific_workbooks": self._prepare_source_specific_workbooks,
            "catalog_and_ledger": self._prepare_catalog_and_ledger,
            "ambiguous_budgets": self._prepare_ambiguous_budgets,
        }
        try:
            builders[fixture]()
        except KeyError as error:
            raise ValueError(f"Unknown RAG regression fixture: {fixture}") from error

    def _prepare_multi_tab_inventory(self) -> None:
        """Create records split across two visible workbook tabs."""
        self.upload_workbook("inventory.xlsx", [
            ("North", [["Item", "Units"], ["N-1", 4], ["N-2", 6]]),
            ("South", [["Item", "Units"], ["S-1", 5]]),
        ])

    def _prepare_people_profiles(self) -> None:
        """Create generic person sheets for exact-name and follow-up patterns."""
        self.upload_workbook("people-profiles.xlsx", [
            ("Asha", [["Role", "Product Designer"], ["Primary Skills", "Research; Prototyping"]]),
            ("Ravi", [["Role", "Data Engineer"], ["Primary Skills", "Python; SQL"]]),
            ("Mei", [["Role", "QA Analyst"], ["Primary Skills", "Test planning; Automation"]]),
        ])

    def _prepare_project_roster(self) -> None:
        """Create filtered count rows that can seed list-style follow-ups."""
        self.upload_workbook("project-roster.xlsx", [
            ("Projects", [["Project", "Status"], ["P-100", "active"], ["P-200", "active"], ["P-300", "closed"]]),
        ])

    def _prepare_campaign_metrics(self) -> None:
        """Create a generic percentage column with a deterministic average."""
        self.upload_workbook("campaign-metrics.xlsx", [
            ("Metrics", [["Campaign", "Completion Rate"], ["Launch A", "50%"], ["Launch B", "100%"]]),
        ])

    def _prepare_source_specific_workbooks(self) -> None:
        """Create two unrelated workbooks so a named source must win."""
        self.upload_workbook("attendance-summary.xlsx", [
            ("Attendance", [["Person", "Present Days"], ["Asha", 20], ["Ravi", 18]]),
        ])
        self.upload_workbook("inspection-report.xlsx", [
            ("Inspection", [["Component", "Rejected"], ["Gear", 5]]),
        ])

    def _prepare_catalog_and_ledger(self) -> None:
        """Create multiple workbooks where only one schema answers the query."""
        self.upload_workbook("catalog.xlsx", [
            ("Catalog", [["Code", "Label"], ["X1", "Alpha"]]),
        ])
        self.upload_workbook("ledger.xlsx", [
            ("Ledger", [["Period", "Amount"], ["March", 25]]),
        ])

    def _prepare_ambiguous_budgets(self) -> None:
        """Create equally plausible workbooks that require user selection."""
        self.upload_workbook("team-a-budget.xlsx", [
            ("Budget", [["Category", "Amount"], ["Travel", 10]]),
        ])
        self.upload_workbook("team-b-budget.xlsx", [
            ("Budget", [["Category", "Amount"], ["Travel", 20]]),
        ])

    def seed_prior_turns(self, case: RagRegressionCase) -> None:
        """Execute prior turns through production orchestration using local fixtures."""
        for turn in case.prior_turns:
            with patch.object(rag_service, "search_chunks", return_value=[]), patch.object(
                rag_service,
                "generate_answer",
                side_effect=AssertionError("Prior-turn fixtures must not call an external generator."),
            ), patch.object(
                rag_service, "log_audit_event"
            ):
                result = rag_service.answer_question(
                    turn.query,
                    1,
                    conversation_id=self.conversation_id,
                )
            if not result.get("grounded"):
                raise AssertionError(f"Prior turn did not establish context: {turn.query}")
            answer = str(result.get("answer") or "").casefold()
            for fact in turn.expected_facts:
                if fact.casefold() not in answer:
                    raise AssertionError(f"Prior turn omitted expected fact {fact!r}: {turn.query}")
            filenames = {
                str(source.get("filename") or "")
                for source in result.get("sources") or []
                if isinstance(source, dict)
            }
            if filenames != set(turn.expected_documents):
                raise AssertionError(
                    f"Prior turn used {sorted(filenames)}, expected {sorted(turn.expected_documents)}"
                )

    @staticmethod
    def _chunks_from_sources(sources: list[dict[str, Any]]) -> tuple[RetrievedChunk, ...]:
        """Normalize app source dictionaries for shared retrieval assertions."""
        return tuple(
            RetrievedChunk(
                chunk_id=(int(source["chunk_id"]) if source.get("chunk_id") is not None else None),
                document_id=(
                    int(source["document_id"]) if source.get("document_id") is not None else None
                ),
                version_id=(
                    int(source["version_id"]) if source.get("version_id") is not None else None
                ),
                filename=str(source.get("filename") or ""),
                text=str(source.get("text") or source.get("content") or source.get("answer") or ""),
                score=(float(source["score"]) if source.get("score") is not None else None),
            )
            for source in sources
        )

    @staticmethod
    def _sheets_from_sources(sources: list[dict[str, Any]]) -> tuple[str, ...]:
        """Extract stable worksheet provenance from structured citations."""
        names = {
            str(location["sheet_name"])
            for source in sources
            if isinstance(source, dict)
            for location in [source.get("source_location")]
            if isinstance(location, dict) and location.get("sheet_name")
        }
        return tuple(sorted(names))

    def observe_retrieval(self, case: RagRegressionCase) -> RetrievalRun:
        """Evaluate routing/source behavior without checking final answer text."""
        self.seed_prior_turns(case)
        if case.prior_turns:
            follow_up = resolve_follow_up(
                owner_id=1,
                conversation_id=self.conversation_id,
                question=case.query,
            )
            if follow_up is None:
                return RetrievalRun(RetrievalObservation(route="unavailable", unavailable=True), ())
            sources = [item for item in follow_up.get("sources") or [] if isinstance(item, dict)]
            return RetrievalRun(
                RetrievalObservation(route="follow_up", chunks=self._chunks_from_sources(sources)),
                self._sheets_from_sources(sources),
            )

        if case.expected_retrieval.forbid_global_search:
            with patch.object(rag_service, "search_chunks", return_value=[]) as search, patch.object(
                rag_service, "log_audit_event"
            ):
                result = rag_service.answer_question(
                    case.query,
                    1,
                    conversation_id=self.conversation_id,
                )
            sources = [item for item in result.get("sources") or [] if isinstance(item, dict)]
            route = "unavailable" if not result.get("grounded") else str(result.get("question_type"))
            return RetrievalRun(
                RetrievalObservation(
                    route=route,
                    chunks=self._chunks_from_sources(sources),
                    unavailable=not bool(result.get("grounded")),
                ),
                self._sheets_from_sources(sources),
                global_search_attempted=search.called,
            )

        structured_available = has_structured_workbook(1)
        structured_requested = structured_available and (
            is_analytical_question(case.query)
            or is_structured_lookup_question(case.query, 1)
        )
        selection = select_sources(
            question=case.query,
            owner_id=1,
            structured_requested=structured_requested,
            searcher=lambda *args, **kwargs: [],
        )
        if selection.path != "structured" or selection.document_id is None:
            route = "unstructured" if selection.path == "retrieval" else selection.path
            return RetrievalRun(
                RetrievalObservation(
                    route=route,
                    selected_document_id=selection.document_id,
                    unavailable=selection.path == "unavailable",
                ),
                (),
            )
        result = analyze_workbook_question(case.query, 1, document_id=selection.document_id)
        sources = [item for item in result.get("sources") or [] if isinstance(item, dict)]
        return RetrievalRun(
            RetrievalObservation(
                route="structured",
                chunks=self._chunks_from_sources(sources),
                selected_document_id=selection.document_id,
                unavailable=not bool(result.get("grounded")),
            ),
            self._sheets_from_sources(sources),
        )

    def observe_answer(self, case: RagRegressionCase) -> dict[str, Any]:
        """Evaluate final answer behavior with external generation disabled."""
        self.seed_prior_turns(case)
        if case.prior_turns:
            result = resolve_follow_up(
                owner_id=1,
                conversation_id=self.conversation_id,
                question=case.query,
            )
            return result or {"answer": UNAVAILABLE_ANSWER, "grounded": False, "sources": []}
        expected_documents = case.expected_retrieval.documents
        if case.expected_retrieval.route == "structured" and len(expected_documents) == 1:
            # Bypass source selection here: retrieval has its own independently
            # generated test, while this path checks deterministic answer behavior.
            document_id = self.uploaded[expected_documents[0]]
            return analyze_workbook_question(case.query, 1, document_id=document_id)
        with patch.object(rag_service, "search_chunks", return_value=[]), patch.object(
            rag_service,
            "generate_answer",
            side_effect=AssertionError("Regression fixtures must not call an external generator."),
        ), patch.object(rag_service, "log_audit_event"):
            return rag_service.answer_question(
                case.query,
                1,
                conversation_id=self.conversation_id,
            )


def assert_regression_retrieval(
    test: unittest.TestCase,
    case: RagRegressionCase,
    run: RetrievalRun,
) -> None:
    """Assert routing and provenance without asserting answer wording."""
    expected = case.expected_retrieval
    test.assertEqual(run.observation.route, expected.route)
    filenames = {chunk.filename for chunk in run.observation.chunks if chunk.filename}
    if not filenames and run.observation.selected_document_id is not None:
        filenames = {
            filename
            for filename, document_id in getattr(test, "harness").uploaded.items()
            if document_id == run.observation.selected_document_id
        }
    test.assertEqual(filenames, set(expected.documents))
    if expected.sheets:
        test.assertEqual(set(run.sheets), set(expected.sheets))
    if expected.forbid_global_search:
        test.assertFalse(run.global_search_attempted, "Follow-up without context searched globally.")


def assert_regression_answer(
    test: unittest.TestCase,
    case: RagRegressionCase,
    result: dict[str, Any],
) -> None:
    """Assert final grounding and facts independently from retrieval assertions."""
    expected = case.expected_answer
    test.assertEqual(bool(result.get("grounded")), expected.grounded)
    answer = str(result.get("answer") or "")
    if expected.unavailable:
        test.assertEqual(answer, UNAVAILABLE_ANSWER)
        test.assertEqual(result.get("sources") or [], [])
    for fact in expected.facts:
        normalized_fact = fact.casefold()
        normalized_answer = answer.casefold()
        if re.fullmatch(r"[+-]?\d+(?:\.\d+)?%?", normalized_fact):
            pattern = rf"(?<![\d.]){re.escape(normalized_fact)}(?!\d)"
            test.assertRegex(normalized_answer, pattern)
        else:
            test.assertIn(normalized_fact, normalized_answer)
    for fact in expected.excluded_facts:
        test.assertNotIn(fact.casefold(), answer.casefold())
    if expected.question_type is not None:
        test.assertEqual(result.get("question_type"), expected.question_type)
