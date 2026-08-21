"""Domain-neutral, schema-driven workbook RAG tests."""

from __future__ import annotations

import asyncio
from contextlib import ExitStack
from io import BytesIO
from pathlib import Path
import tempfile
from types import SimpleNamespace
import unittest
from unittest.mock import patch

from fastapi import UploadFile
from openpyxl import Workbook
from starlette.requests import Request

from app import database
from app.routes import upload
from app.services import structured_ingestion, vector_search, vector_store
from app.services.chat_context import save_grounded_context
from app.services.rag_service import answer_question
from app.services.source_selection import select_sources
from app.services.vector_store import reset_vector_store_for_tests
from app.services.workbook_analysis import (
    RowRecord,
    analyze_workbook_question,
    _column_score,
    _row_filters,
)


ORG = "00000000-0000-4000-8000-000000000001"


def _request() -> Request:
    return Request({
        "type": "http",
        "method": "POST",
        "path": "/documents/upload",
        "headers": [],
        "client": ("test", 1),
    })


def _workbook_bytes(sheets: list[tuple[str, list[list[object]], str]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows, state in sheets:
        sheet = workbook.create_sheet(name)
        sheet.sheet_state = state
        for row in rows:
            sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _merged_attendance_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Monthly Attendance Report"
    sheet.merge_cells("A3:A4")
    sheet["A3"] = "Employee ID"
    sheet.merge_cells("B3:B4")
    sheet["B3"] = "Employee Name"
    sheet.merge_cells("C3:D3")
    sheet["C3"] = "Attendance"
    sheet["C4"] = "Present Days"
    sheet["D4"] = "Absent Days"
    sheet.append(["E001", "Aparna", 20, 2])
    sheet.append(["E002", "Hari", 18, 4])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class _FakePdfPage:
    """Test double that exposes plain and layout pypdf extraction modes."""

    def __init__(self, plain: str, layout: str | None = None) -> None:
        self.plain = plain
        self.layout = layout or plain

    def extract_text(self, *args, **kwargs) -> str:
        if kwargs.get("visitor_text"):
            return self.plain
        if kwargs.get("extraction_mode") == "layout":
            return self.layout
        return self.plain


class WorkbookDomainNeutralTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        root = Path(self.temporary.name)
        self.database_path = root / "test.db"
        self.upload_path = root / "uploads"
        self.stack = ExitStack()
        self.stack.enter_context(patch.object(database, "DATABASE_PATH", self.database_path))
        self.stack.enter_context(patch.object(database, "UPLOAD_DIRECTORY", self.upload_path))
        self.stack.enter_context(patch.object(vector_store.settings, "vector_store", "sqlite"))
        self.stack.enter_context(patch.object(vector_store.settings, "vector_store_provider", "sqlite"))
        self.stack.enter_context(patch.object(vector_store.settings, "qdrant_local_path", ""))
        self.stack.enter_context(patch.object(upload, "UPLOAD_DIRECTORY", self.upload_path))
        self.stack.enter_context(patch.object(upload, "enforce_request_limit", lambda *args, **kwargs: None))
        self.stack.enter_context(patch.object(upload, "log_audit_event", lambda **kwargs: None))
        self.stack.enter_context(
            patch.object(
                upload,
                "create_embeddings",
                lambda chunks: [
                    ([1.0, 0.0] if "FINAL_ONLY" in chunk else [0.0, 1.0]) + [0.0] * 382
                    for chunk in chunks
                ],
            )
        )
        database.initialize_database()
        reset_vector_store_for_tests()
        with database.get_connection() as connection:
            connection.executemany(
                "INSERT INTO users (id, email, password_hash) VALUES (?, ?, 'hash')",
                [(1, "owner@example.com"), (2, "other@example.com")],
            )

    def tearDown(self) -> None:
        self.stack.close()
        self.temporary.cleanup()

    def upload_workbook(
        self,
        sheets: list[tuple[str, list[list[object]], str]],
        filename: str,
        owner_id: int = 1,
    ) -> dict[str, object]:
        file = UploadFile(file=BytesIO(_workbook_bytes(sheets)), filename=filename)
        return asyncio.run(upload._process_document_upload(_request(), file, {"id": owner_id}))

    def upload_text(self, text: str, filename: str, owner_id: int = 1) -> dict[str, object]:
        file = UploadFile(file=BytesIO(text.encode("utf-8")), filename=filename)
        return asyncio.run(upload._process_document_upload(_request(), file, {"id": owner_id}))

    def upload_pdf_table(self, text: str, filename: str = "metrics.pdf") -> dict[str, object]:
        page = _FakePdfPage(text)
        file = UploadFile(file=BytesIO(b"%PDF-1.4\n%%EOF"), filename=filename)
        with patch("pypdf.PdfReader", return_value=SimpleNamespace(pages=[page])):
            return asyncio.run(upload._process_document_upload(_request(), file, {"id": 1}))

    def upload_pdf_layout_table(self, plain: str, layout: str, filename: str = "final-inspection-rejection.pdf") -> dict[str, object]:
        page = _FakePdfPage(plain, layout)
        file = UploadFile(file=BytesIO(b"%PDF-1.4\n%%EOF"), filename=filename)
        with patch("pypdf.PdfReader", return_value=SimpleNamespace(pages=[page])):
            return asyncio.run(upload._process_document_upload(_request(), file, {"id": 1}))

    def upload_merged_attendance_workbook(self) -> dict[str, object]:
        file = UploadFile(file=BytesIO(_merged_attendance_bytes()), filename="attendance.xlsx")
        return asyncio.run(upload._process_document_upload(_request(), file, {"id": 1}))

    def test_multi_sheet_tables_are_schema_indexed_and_counted(self) -> None:
        result = self.upload_workbook(
            [
                ("North", [["Item", "Units"], ["A-1", 4], ["A-2", 6]], "visible"),
                ("South", [["Item", "Units"], ["B-1", 5]], "visible"),
            ],
            "inventory.xlsx",
        )

        answer = analyze_workbook_question("How many records are there?", 1, document_id=int(result["document_id"]))

        self.assertEqual(len(result["workbook"]["processed_sheets"]), 2)
        self.assertIn("Count: 3", answer["answer"])
        self.assertEqual({source["source_location"]["sheet_name"] for source in answer["sources"]}, {"North", "South"})

    def test_filtered_total_uses_all_rows_not_vector_excerpt(self) -> None:
        result = self.upload_workbook(
            [
                ("Q1", [["Period", "Region", "Amount"], ["2026-01-01", "East", 10], ["2026-02-01", "East", 20]], "visible"),
                ("Q2", [["Period", "Region", "Amount"], ["2026-07-01", "East", 30], ["2026-07-15", "West", 40]], "visible"),
            ],
            "ledger.xlsx",
        )

        answer = analyze_workbook_question("What is the total amount for July?", 1, document_id=int(result["document_id"]))

        self.assertIn("70", answer["answer"])
        self.assertNotIn("100", answer["answer"])
        self.assertEqual(answer["sources"][0]["source_location"]["sheet_name"], "Q2")

    def test_automatic_document_selection_uses_schema_and_values(self) -> None:
        first = self.upload_workbook(
            [("Catalog", [["Code", "Label"], ["X1", "Alpha"]], "visible")],
            "catalog.xlsx",
        )
        second = self.upload_workbook(
            [("Ledger", [["Period", "Amount"], ["March", 25]], "visible")],
            "ledger.xlsx",
        )

        answer = analyze_workbook_question("What is the total amount for March?", 1)

        self.assertIn("25", answer["answer"])
        self.assertEqual(answer["sources"][0]["document_id"], int(second["document_id"]))
        self.assertNotEqual(answer["sources"][0]["document_id"], int(first["document_id"]))

    def test_common_short_words_are_not_cell_value_filters(self) -> None:
        rows = [
            RowRecord("Sheet1", 2, {"State": "IN", "Amount": 5}),
            RowRecord("Sheet1", 3, {"State": "OUT", "Amount": 7}),
        ]

        self.assertEqual(_row_filters(rows, "What is the total amount in July?"), {})
        self.assertEqual(_row_filters(rows, "Show state IN"), {"State": {"in"}})

    def test_single_character_and_substring_headers_do_not_create_relevance(self) -> None:
        self.assertEqual(_column_score("A", "How many products rejected?"), 0)
        self.assertEqual(_column_score("art", "What is the partial count?"), 0)
        self.assertGreater(_column_score("Product Count", "How many product records?"), 0)

    def test_unrelated_workbook_cannot_win_generic_count(self) -> None:
        self.upload_workbook(
            [("Log", [["A", "Column 2"], ["x", "present"], ["y", "absent"]], "visible")],
            "neutral.xlsx",
        )

        answer = analyze_workbook_question("How many products rejected?", 1)

        self.assertFalse(answer["grounded"])
        self.assertEqual(answer["sources"], [])

    def test_relevant_pdf_style_retrieval_wins_over_unrelated_workbook(self) -> None:
        self.upload_workbook(
            [("Log", [["A", "Column 2"], ["x", "present"], ["y", "absent"]], "visible")],
            "neutral.xlsx",
        )
        text = self.upload_text(
            "The uploaded file discusses product rejection evidence.",
            "quality-summary.txt",
        )
        with database.get_connection() as connection:
            version_id = connection.execute(
                "SELECT current_version_id FROM documents WHERE id = ?",
                (int(text["document_id"]),),
            ).fetchone()["current_version_id"]
        pdf_source = {
            "document_id": int(text["document_id"]),
            "version_id": int(version_id),
            "filename": "quality-summary.txt",
            "content": "The uploaded file discusses product rejection evidence.",
            "source_type": "text",
            "source_location": {"section": "body"},
            "score": 0.82,
        }

        with patch("app.services.rag_service.search_chunks", return_value=[pdf_source]), patch(
            "app.services.rag_service.generate_answer",
            return_value={
                "answer": "The file contains product rejection evidence (quality-summary.pdf, page 2).",
                "prompt_tokens": 100,
                "completion_tokens": 20,
            },
        ), patch("app.services.rag_service.reserve_groq_call"), patch(
            "app.services.rag_service.record_groq_tokens"
        ), patch("app.services.rag_service.log_audit_event"):
            answer = answer_question("What are the product rejections?", 1)

        self.assertEqual(answer["question_type"], "retrieval")
        self.assertEqual(answer["sources"][0]["filename"], "quality-summary.txt")

    def test_quantity_question_uses_sum_when_numeric_column_matches(self) -> None:
        result = self.upload_workbook(
            [("Stock", [["Item", "Units"], ["A", 2], ["B", 3]], "visible")],
            "stock.xlsx",
        )

        answer = analyze_workbook_question(
            "How many units are there?",
            1,
            document_id=int(result["document_id"]),
        )

        self.assertIn("Total Units: 5", answer["answer"])

    def test_status_filtering_uses_filtered_count(self) -> None:
        result = self.upload_workbook(
            [("Cases", [["Item", "Status"], ["A", "open"], ["B", "closed"]], "visible")],
            "cases.xlsx",
        )

        answer = analyze_workbook_question(
            "How many open items?",
            1,
            document_id=int(result["document_id"]),
        )

        self.assertIn("Count: 1", answer["answer"])
        self.assertEqual(answer["_context"]["filters"], {"Status": ["open"]})

    def test_grouped_result_is_deterministic_structured_evidence(self) -> None:
        result = self.upload_workbook(
            [("Data", [["Category", "Amount"], ["A", 3], ["B", 5], ["A", 7]], "visible")],
            "grouped.xlsx",
        )

        answer = analyze_workbook_question("Group amount by category", 1, document_id=int(result["document_id"]))

        self.assertIn("A: 10", answer["answer"])
        self.assertIn("B: 5", answer["answer"])
        self.assertTrue(answer["grounded"])

    def test_pdf_table_is_structured_for_count_total_and_percentage(self) -> None:
        result = self.upload_pdf_table(
            "\n".join([
                "Defect | Quantity | Rate",
                "Crack | 2 | 4.0%",
                "Scratch | 3 | 6.0%",
            ])
        )
        document_id = int(result["document_id"])

        count = analyze_workbook_question("How many defects are found?", 1, document_id=document_id)
        total = analyze_workbook_question("What is the total quantity?", 1, document_id=document_id)
        rate = analyze_workbook_question("What is the overall rate?", 1, document_id=document_id)

        self.assertIn("Count: 2", count["answer"])
        self.assertIn("Total Quantity: 5", total["answer"])
        self.assertIn("Average Rate: 5", rate["answer"])
        self.assertEqual(total["sources"][0]["source_type"], "pdf")
        self.assertEqual(total["sources"][0]["source_location"]["page_start"], 1)
        self.assertEqual(total["sources"][0]["source_location"]["table_name"], "Table 1")

    def test_flattened_pdf_uses_layout_table_for_rejection_counts(self) -> None:
        result = self.upload_pdf_layout_table(
            "FINAL INSPECTION REJECTION Component Rejection Count Gear 5 Valve 3",
            "\n".join([
                "FINAL INSPECTION REJECTION",
                "Component  Rejection Count",
                "Gear       5",
                "Valve      3",
            ]),
        )
        document_id = int(result["document_id"])

        total = analyze_workbook_question(
            "What is the total rejection count?",
            1,
            document_id=document_id,
        )
        filtered = analyze_workbook_question(
            "What is the rejection count for Gear?",
            1,
            document_id=document_id,
        )

        self.assertIn("Total Rejection Count: 8", total["answer"])
        self.assertIn("Total Rejection Count: 5", filtered["answer"])
        self.assertEqual(total["sources"][0]["source_type"], "pdf")

    def test_merged_attendance_headers_are_preserved(self) -> None:
        result = self.upload_merged_attendance_workbook()
        document_id = int(result["document_id"])

        with database.get_connection() as connection:
            sheet = connection.execute(
                """SELECT ws.headers_json
                   FROM workbook_sheets ws
                   JOIN documents d ON d.content_id = ws.content_id
                   WHERE d.id = ?""",
                (document_id,),
            ).fetchone()
        answer = analyze_workbook_question(
            "What is the total present days for Aparna in attendance?",
            1,
            document_id=document_id,
        )

        self.assertIn("Employee ID", sheet["headers_json"])
        self.assertIn("Employee Name", sheet["headers_json"])
        self.assertNotIn("Column 1", sheet["headers_json"])
        self.assertNotIn("Column 2", sheet["headers_json"])
        self.assertIn("Total Present Days: 20", answer["answer"])

    def test_attendance_in_time_question_returns_in_rows(self) -> None:
        result = self.upload_workbook(
            [
                (
                    "Sheet1",
                    [
                        ["EmpLoyee Name & No", "Attendance Direction", "1st", "2nd"],
                        ["Aparna", "IN", "09:10:00", "09:15:00"],
                        [101, "OUT", "18:10:00", "18:20:00"],
                        ["Hari", "IN", "10:00:00", "10:05:00"],
                        [102, "OUT", "19:00:00", "19:05:00"],
                    ],
                    "visible",
                )
            ],
            "Attentence.xlsx",
        )

        answer = analyze_workbook_question(
            "What is the in time of all employees in atendence sheet?",
            1,
            document_id=int(result["document_id"]),
        )

        self.assertEqual(answer["question_type"], "structured_analysis")
        self.assertTrue(answer["grounded"])
        self.assertIn("Matching records (2)", answer["answer"])
        self.assertIn("Aparna", answer["answer"])
        self.assertIn("Hari", answer["answer"])
        self.assertIn("09:10:00", answer["answer"])
        self.assertNotIn("18:10:00", answer["answer"])
        self.assertEqual(answer["_context"]["filters"], {"Attendance Direction": ["in"]})

    def test_attendance_filter_wins_over_unrelated_semantic_match(self) -> None:
        attendance = self.upload_workbook(
            [
                (
                    "Sheet1",
                    [
                        ["EmpLoyee Name & No", "Attendance Direction", "1st"],
                        ["Aparna", "IN", "09:10:00"],
                        [101, "OUT", "18:10:00"],
                    ],
                    "visible",
                )
            ],
            "Attentence.xlsx",
        )
        tracker = self.upload_text(
            "Task tracker about employees and sheet details repeated employees sheet.",
            "aparna_task_tracker.txt",
        )

        with database.get_connection() as connection:
            tracker_version = connection.execute(
                "SELECT current_version_id FROM documents WHERE id = ?",
                (int(tracker["document_id"]),),
            ).fetchone()["current_version_id"]

        noisy_source = {
            "document_id": int(tracker["document_id"]),
            "version_id": int(tracker_version),
            "filename": "aparna_task_tracker.txt",
            "content": "employees sheet " * 20,
            "source_type": "excel",
            "source_location": {"sheet_name": "Tasks", "row_start": 1},
            "score": 0.99,
        }

        with patch("app.services.rag_service.search_chunks", return_value=[noisy_source]), patch(
            "app.services.rag_service.generate_answer",
            side_effect=AssertionError("validated attendance rows should not call the LLM"),
        ):
            answer = answer_question(
                "What is the in time of all employees in atendence sheet?",
                1,
            )

        self.assertEqual(answer["question_type"], "structured_analysis")
        self.assertIn("09:10:00", answer["answer"])
        self.assertEqual(answer["sources"][0]["document_id"], int(attendance["document_id"]))

    def test_attendance_day_follow_up_uses_prior_row_context(self) -> None:
        self.upload_workbook(
            [
                (
                    "Sheet1",
                    [
                        ["EmpLoyee Name & No", "Attendance Direction", "1st", "18th"],
                        ["D BALAJI", "IN", None, "22:00:00"],
                        [33, "OUT", None, "08:25:00"],
                    ],
                    "visible",
                )
            ],
            "Attentence.xlsx",
        )

        first = answer_question("D BALAJI in time", 1, conversation_id="attendance-chat")
        follow_up = answer_question("on 18 th", 1, conversation_id="attendance-chat")

        self.assertTrue(first["grounded"])
        self.assertEqual(follow_up["question_type"], "follow_up")
        self.assertIn("18th from the prior grounded result", follow_up["answer"])
        self.assertIn("D BALAJI: 22:00:00", follow_up["answer"])
        self.assertNotIn("08:25:00", follow_up["answer"])

    def test_person_sheet_career_goal_lookup_handles_mixed_layouts(self) -> None:
        sandhiya = "To grow as a strong UI/UX Product Designer with React knowledge."
        bathmaraj = "Deepen expertise in DevOps and infrastructure automation."
        result = self.upload_workbook(
            [
                (
                    "Sandhiya",
                    [
                        ["Role", "UI/UX Designer", "Pick from list"],
                        ["Career goals", sandhiya, "Prompt"],
                    ],
                    "visible",
                ),
                (
                    "Bathmaraj",
                    [
                        ["Name", "Bathmaraj V"],
                        ["Career goals", bathmaraj],
                    ],
                    "visible",
                ),
            ],
            "Employee Skill Matrix.xlsx",
        )

        answer = answer_question(
            "sandhiyas career goal",
            1,
            document_id=int(result["document_id"]),
        )

        self.assertTrue(answer["grounded"])
        self.assertEqual(answer["question_type"], "structured_analysis")
        self.assertIn(sandhiya, answer["answer"])
        self.assertNotIn(bathmaraj, answer["answer"])
        self.assertEqual(answer["sources"][0]["source_location"]["sheet_name"], "Sandhiya")

    def test_form_style_field_lookup_returns_direct_value_without_source_column(self) -> None:
        responsibilities = "Gather requirements and prepare BRD, FRD, and process documents."
        result = self.upload_workbook(
            [
                (
                    "Keerthana",
                    [
                        ["Column1", "Role", "Business Analyst", "Pick from list"],
                        [
                            None,
                            "Expected responsibilities",
                            responsibilities,
                            "What are your current responsibility and roles in details?",
                        ],
                    ],
                    "visible",
                )
            ],
            "Employee Skill Matrix (1).xlsx",
        )

        answer = answer_question(
            "expected responsibilities of keerthana",
            1,
            document_id=int(result["document_id"]),
        )

        self.assertTrue(answer["grounded"])
        self.assertEqual(answer["question_type"], "structured_analysis")
        self.assertIn("Keerthana", answer["answer"])
        self.assertIn("Role: Business Analyst", answer["answer"])
        self.assertIn(responsibilities, answer["answer"])
        self.assertNotIn("Matching records", answer["answer"])
        self.assertNotIn("Source", answer["answer"])
        self.assertEqual(answer["sources"][0]["source_location"]["sheet_name"], "Keerthana")

        with database.get_connection() as connection:
            content = connection.execute(
                "SELECT content_id FROM documents WHERE id = ?",
                (int(result["document_id"]),),
            ).fetchone()
            row = connection.execute(
                """SELECT ws.headers_json, wr.values_json
                   FROM workbook_sheets ws
                   JOIN workbook_rows wr ON wr.sheet_id = ws.id
                   WHERE ws.content_id = ?""",
                (int(content["content_id"]),),
            ).fetchone()
        self.assertIn("Employee", row["headers_json"])
        self.assertIn("Expected Responsibilities", row["headers_json"])
        self.assertNotIn("Column 1", row["headers_json"])
        self.assertNotIn("Pick from list", row["values_json"])

    def test_employee_skill_comparison_uses_all_profile_sheets(self) -> None:
        result = self.upload_workbook(
            [
                ("Keerthana", [["Role", "Business Analyst", "Pick from list"], ["Experience Level", "Junior"], ["Primary Skills", "BRD/FRD * UAT coordination"], ["Secondary Skills", "SQL"]], "visible"),
                ("Sandhiya", [["Role", "UI/UX Designer", "Pick from list"], ["Experience Level", "Mid"], ["Primary Skills", "Wireframes; React"], ["Secondary Skills", "Figma"]], "visible"),
                ("Kirubba", [["Role", "Developer"], ["Experience Level", "Junior"], ["Primary Skills", "Python"], ["Secondary Skills", "FastAPI"]], "visible"),
                ("Bathmaraj", [["Role", "DevOps Engineer"], ["Experience Level", "Senior"], ["Primary Skills", "CI/CD"], ["Secondary Skills", "Azure"]], "visible"),
                ("Aparna", [["Role", "QA Analyst"], ["Experience Level", "Mid"], ["Primary Skills", "Test planning"], ["Secondary Skills", "Automation"]], "visible"),
                ("Hari", [["Role", "Data Analyst"], ["Experience Level", "Senior"], ["Primary Skills", "Dashboards"], ["Secondary Skills", "SQL"]], "visible"),
                ("Nandhini", [["Role", "QA Associate"], ["Experience Level", "Mid"], ["Primary Skills", "1. Manual testing 2.Regression testing 3.5+ years in QA"], ["Secondary Skills", "1.Mentoring 2. Knowledge sharing"]], "visible"),
            ],
            "Employee Skill Matrix (1).xlsx",
        )

        answer = answer_question(
            "compare skills of all employees in skill matrix",
            1,
            document_id=int(result["document_id"]),
            conversation_id="skill-chat",
        )

        self.assertTrue(answer["grounded"])
        self.assertEqual(answer["question_type"], "structured_analysis")
        for name in ("Keerthana", "Sandhiya", "Kirubba", "Bathmaraj", "Aparna", "Hari", "Nandhini"):
            self.assertIn(name, answer["answer"])
        self.assertIn("1. **Keerthana**", answer["answer"])
        self.assertIn("7. **Nandhini**", answer["answer"])
        self.assertIn("**Primary skills**", answer["answer"])
        self.assertIn("**Secondary skills**", answer["answer"])
        self.assertIn("- BRD/FRD", answer["answer"])
        self.assertIn("- UAT coordination", answer["answer"])
        self.assertIn("- Manual testing", answer["answer"])
        self.assertIn("- Regression testing", answer["answer"])
        self.assertIn("- Regression testing 3.5+ years in QA", answer["answer"])
        self.assertNotIn("- 5+ years in QA", answer["answer"])
        self.assertNotRegex(answer["answer"], r"-\s+\d+\.")
        self.assertNotIn("1. Manual testing 2.Regression testing", answer["answer"])
        self.assertNotIn(" more", answer["answer"])
        self.assertNotIn("| Employee | Role | Experience | Primary skills | Secondary skills |", answer["answer"])
        self.assertNotIn("Role:", answer["answer"])
        self.assertNotIn("Experience:", answer["answer"])

    def test_employee_name_fuzzy_matching_and_follow_up_stay_structured(self) -> None:
        result = self.upload_workbook(
            [
                ("Sandhiya", [["Role", "UI/UX Designer"], ["Primary Skills", "Wireframes"]], "visible"),
                ("Kirubba", [["Role", "Developer"], ["Primary Skills", "Python"]], "visible"),
                ("Bathmaraj", [["Role", "DevOps Engineer"], ["Primary Skills", "CI/CD"]], "visible"),
                ("Aparna", [["Role", "QA Analyst"], ["Primary Skills", "Test planning"]], "visible"),
            ],
            "Employee Skill Matrix (1).xlsx",
        )

        first = answer_question(
            "compare skills of Sandhiya and Aparna",
            1,
            document_id=int(result["document_id"]),
            conversation_id="employee-followup",
        )
        follow_up = answer_question(
            "what about Sandhiya, Kiruba, Bathmaraj?",
            1,
            conversation_id="employee-followup",
        )

        self.assertTrue(first["grounded"])
        self.assertTrue(follow_up["grounded"])
        self.assertEqual(follow_up["question_type"], "follow_up")
        self.assertIn("1. **Sandhiya**", follow_up["answer"])
        self.assertIn("Kirubba", follow_up["answer"])
        self.assertIn("Bathmaraj", follow_up["answer"])
        self.assertIn("**Primary skills**", follow_up["answer"])
        self.assertNotIn("Aparna", follow_up["answer"])
        self.assertNotIn("| Employee | Role | Experience |", follow_up["answer"])
        self.assertNotIn("Role:", follow_up["answer"])

    def test_named_document_routes_before_mixed_semantic_search(self) -> None:
        attendance = self.upload_merged_attendance_workbook()
        self.upload_workbook(
            [("Inspection", [["Component", "Rejection Count"], ["Gear", 5]], "visible")],
            "final-inspection-rejection.xlsx",
        )
        requested_document_ids: list[int | None] = []

        def searcher(*args, **kwargs):
            requested_document_ids.append(kwargs.get("document_id"))
            document_id = int(kwargs["document_id"])
            return [{
                "document_id": document_id,
                "version_id": 1,
                "filename": "attendance.xlsx",
                "content": "Attendance evidence",
                "source_type": "excel",
                "source_location": {},
                "score": 0.9,
            }]

        result = select_sources(
            question="What does the attendance upload say?",
            owner_id=1,
            searcher=searcher,
        )

        self.assertEqual(requested_document_ids[0], int(attendance["document_id"]))
        self.assertEqual(result.document_id, int(attendance["document_id"]))

    def test_ambiguous_numeric_target_returns_unavailable(self) -> None:
        result = self.upload_workbook(
            [("Data", [["First Amount", "Second Amount"], [10, 20]], "visible")],
            "ambiguous.xlsx",
        )

        answer = analyze_workbook_question(
            "What is the total amount?",
            1,
            document_id=int(result["document_id"]),
        )

        self.assertFalse(answer["grounded"])
        self.assertEqual(answer["sources"], [])

    def test_complete_structured_answer_wins_over_partial_vector_context(self) -> None:
        result = self.upload_workbook(
            [("Data", [["Period", "Amount"], ["April", 5], ["April", 15]], "visible")],
            "complete.xlsx",
        )
        partial = {
            "document_id": int(result["document_id"]),
            "version_id": 0,
            "filename": "complete.xlsx",
            "content": "Period: April\nAmount: 5",
            "source_type": "excel",
            "source_location": {"sheet_name": "Data", "row_start": 2, "row_end": 2},
            "score": 0.99,
        }

        with patch("app.services.rag_service.search_chunks", return_value=[partial]):
            answer = answer_question("What is the total amount for April?", 1)

        self.assertIn("20", answer["answer"])
        self.assertEqual(answer["question_type"], "structured_analysis")

    def test_low_score_retrieval_requires_overlap_and_keeps_sources_empty_when_unavailable(self) -> None:
        unrelated = {
            "document_id": 50,
            "version_id": 51,
            "filename": "notes.txt",
            "content": "Completely different content.",
            "source_type": "text",
            "source_location": {},
            "score": 0.1,
        }

        with patch("app.services.rag_service.has_structured_workbook", return_value=False), patch(
            "app.services.rag_service.is_analytical_question", return_value=False
        ), patch("app.services.rag_service.search_chunks", side_effect=[[], [unrelated]]), patch(
            "app.services.rag_service.generate_answer", side_effect=AssertionError("unrelated source must not be used")
        ), patch("app.services.rag_service.log_audit_event"):
            answer = answer_question("Find alpha beta", 1)

        self.assertEqual(answer["answer"], "Information not available in the uploaded files.")
        self.assertFalse(answer["grounded"])
        self.assertEqual(answer["sources"], [])

    def test_follow_up_uses_latest_grounded_context_within_same_chat(self) -> None:
        result = self.upload_workbook(
            [("Rows", [["Code", "Amount"], ["R1", 9], ["R2", 11]], "visible")],
            "rows.xlsx",
        )
        total = answer_question("What is the total amount?", 1, conversation_id="chat-a")
        follow_up = answer_question("list them", 1, conversation_id="chat-a")
        other_chat = answer_question("list them", 1, conversation_id="chat-b")
        other_user = answer_question("list them", 2, conversation_id="chat-a")

        self.assertIn("20", total["answer"])
        self.assertIn("9", follow_up["answer"])
        self.assertIn("11", follow_up["answer"])
        self.assertEqual({source["document_id"] for source in follow_up["sources"]}, {int(result["document_id"])})
        self.assertEqual(other_chat["sources"], [])
        self.assertEqual(other_user["sources"], [])

    def test_clear_topic_change_does_not_reuse_stale_context(self) -> None:
        result = self.upload_workbook(
            [("Rows", [["Code", "Amount"], ["R1", 9], ["R2", 11]], "visible")],
            "rows.xlsx",
        )
        answer_question("What is the total amount?", 1, conversation_id="chat-a")

        with patch("app.services.rag_service.search_chunks", return_value=[]), patch(
            "app.services.rag_service.log_audit_event"
        ):
            changed = answer_question("show invoices", 1, conversation_id="chat-a")

        self.assertEqual(int(result["document_id"]), 1)
        self.assertFalse(changed["grounded"])
        self.assertEqual(changed["sources"], [])

    def test_aggregate_citation_carries_complete_row_range(self) -> None:
        result = self.upload_workbook(
            [("Cases", [["Item", "Status"], ["A", "open"], ["B", "open"]], "visible")],
            "cases.xlsx",
        )

        answer = analyze_workbook_question(
            "How many open items?",
            1,
            document_id=int(result["document_id"]),
        )

        location = answer["sources"][0]["source_location"]
        self.assertEqual(location["row_start"], 2)
        self.assertEqual(location["row_end"], 3)

    def test_rating_label_lookup_returns_score_from_exact_row(self) -> None:
        result = self.upload_workbook(
            [
                (
                    "Review",
                    [
                        [None, None, "Review Ratio", None, None, None, "Rating", "Rating Score", "Minimum", "Maximum"],
                        [None, None, "Objective %", 0.5, None, None, "Unsatisfactory", 1, 0, 1.49],
                        [None, None, "Competency %", 0.5, None, None, "Needs Improvement", 2, 1.5, 2.49],
                        [None, None, None, None, None, None, "Meets Expectations", 3, 2.5, 3.49],
                        [None, None, None, None, None, None, "Exceeds Expectations", 4, 3.5, 4.49],
                        [None, None, None, None, None, None, "Exceptional", 5, 4.5, 5],
                        [None, "Objectives", "Rating", "Rating Score", None, None, "Review summary", "AvgRatingScore", "Rating", "Ratio of Rating Score"],
                        [None, "G1", "Meets Expectations", 3, None, None, "Objectives", 3.2, "Meets Expectations", 1.6],
                        [None, "G2", "Exceptional", 5, None, None, "Competency", 1, "Unsatisfactory", 0.5],
                        [None, "G3", "Unsatisfactory", 1, None, None, None, None, None, None],
                        [None, "Competency", "Rating", "Rating Score", None, None, None, None, None, None],
                        [None, "C1", "Unsatisfactory", 1, None, None, None, None, None, None],
                    ],
                    "visible",
                )
            ],
            "Review_Calculation.xlsx",
        )
        document_id = int(result["document_id"])

        objective = analyze_workbook_question("objectives percentage", 1, document_id=document_id)
        unsatisfactory = analyze_workbook_question("unsatisfactory rating", 1, document_id=document_id)
        meets = analyze_workbook_question("give me the meets expectation rating score", 1, document_id=document_id)
        exceeds = analyze_workbook_question("give me the Exceeds Expectations rating score", 1, document_id=document_id)

        self.assertIn("Objective %: 0.5", objective["answer"])
        self.assertEqual(objective["sources"][0]["source_location"]["row_start"], 2)
        self.assertIn("Rating Score: 1", unsatisfactory["answer"])
        self.assertIn("Rating Score: 3", meets["answer"])
        self.assertIn("Rating Score: 4", exceeds["answer"])
        self.assertNotIn("Rating Score: Rating Score", unsatisfactory["answer"])
        self.assertNotIn("Review summary: Objectives", meets["answer"])
        self.assertEqual(meets["sources"][0]["source_location"]["row_start"], 4)
        self.assertEqual(exceeds["sources"][0]["source_location"]["row_start"], 5)

    def test_unavailable_and_acl_isolation_have_no_sources(self) -> None:
        result = self.upload_workbook(
            [("Private", [["Reference", "Value"], ["SECRET", 500]], "visible")],
            "private.xlsx",
        )

        unavailable = analyze_workbook_question("What is the total missing field?", 1, document_id=int(result["document_id"]))
        inaccessible = analyze_workbook_question("What is the total value?", 2, document_id=int(result["document_id"]))

        self.assertEqual(unavailable["sources"], [])
        self.assertFalse(unavailable["grounded"])
        self.assertEqual(inaccessible["sources"], [])
        self.assertNotIn("500", inaccessible["answer"])

    def test_source_plan_mismatch_is_not_saved_as_follow_up_context(self) -> None:
        save_grounded_context(
            owner_id=1,
            conversation_id="mismatch",
            question="older",
            result={
                "answer": "Grounded",
                "grounded": True,
                "sources": [{"document_id": 1, "version_id": 2, "filename": "a.xlsx"}],
                "_context": {"document_ids": [99], "version_ids": [2], "result_type": "count"},
            },
        )

        answer = answer_question("show them", 1, conversation_id="mismatch")

        self.assertEqual(answer["sources"], [])

    def test_invalid_provenance_row_reference_is_not_reused_for_follow_up(self) -> None:
        """Persisted follow-up plans must stay inside their structured provenance range."""
        uploaded = self.upload_workbook(
            [("Ledger", [["Period", "Amount"], ["March", 25]], "visible")],
            "ledger.xlsx",
        )
        result = analyze_workbook_question("What is the total amount for March?", 1, document_id=int(uploaded["document_id"]))
        result["_context"]["row_refs"] = [{
            "document_id": int(uploaded["document_id"]), "sheet": "Ledger", "row_number": 999,
        }]
        save_grounded_context(
            owner_id=1,
            conversation_id="invalid-provenance",
            question="What is the total amount for March?",
            result=result,
        )

        answer = answer_question("show those", 1, conversation_id="invalid-provenance")

        self.assertFalse(answer["grounded"])
        self.assertEqual(answer["sources"], [])

    def test_reindex_preserves_vectors_and_sheet_locations(self) -> None:
        result = self.upload_workbook(
            [("First", [["Code", "Description"], ["A-1", "ordinary"]], "visible"), ("Final", [["Code", "Description"], ["Z-9", "FINAL_ONLY"]], "visible")],
            "projects.xlsx",
        )
        document_id = int(result["document_id"])
        with database.get_connection() as connection:
            before = connection.execute(
                "SELECT vector_point_id FROM chunks WHERE document_id = ? ORDER BY chunk_index",
                (document_id,),
            ).fetchall()
        vectors = {str(row["vector_point_id"]): [0.0, 1.0] + [0.0] * 382 for row in before}

        class RecordingStore:
            def __init__(self) -> None:
                self.batches = []

            def get_vectors(self, point_ids):
                return {point_id: vectors[point_id] for point_id in point_ids if point_id in vectors}

            def upsert_chunks(self, points):
                self.batches.append(points)

        store = RecordingStore()
        with patch.object(structured_ingestion, "get_vector_store", return_value=store), patch.object(
            structured_ingestion,
            "create_embeddings",
            return_value=[[1.0, 0.0] + [0.0] * 382 for _ in range(len(before))],
        ):
            status = structured_ingestion.reindex_existing_spreadsheet_document(
                document_id=document_id,
                owner_id=1,
                organization_id=ORG,
            )

        with patch.object(vector_search, "create_embeddings", return_value=[[1.0, 0.0] + [0.0] * 382]):
            matches = vector_search.search_chunks("FINAL_ONLY", owner_id=1, document_id=document_id, limit=1)

        self.assertEqual(status.status, "completed")
        self.assertEqual(len(store.batches), 1)
        self.assertEqual(matches[0]["sheet_name"], "Final")

    def test_complete_multitab_aggregations_do_not_use_vector_or_llm(self) -> None:
        """All arithmetic uses complete persisted workbook rows across eligible tabs."""
        result = self.upload_workbook(
            [
                ("January", [["Category", "Amount", "Rate"], ["Hardware", 10, "10%"], ["Software", 20, "20%"]], "visible"),
                ("February", [["Category", "Amount", "Rate"], ["Hardware", 30, "30%"], ["Software", 40, "40%"]], "visible"),
            ],
            "metrics.xlsx",
        )
        document_id = int(result["document_id"])

        cases = {
            "How many records?": "Count: 4",
            "How many distinct categories?": "Unique Category: 2",
            "What is the total amount for Hardware?": "Total Amount: 40",
            "What is the average rate?": "Average Rate: 25",
            "What is the minimum amount?": "Minimum Amount: 10",
            "What is the maximum amount?": "Maximum Amount: 40",
            "What is the overall rate?": "Average Rate: 25",
            "Group amount by category": "Hardware: 40",
            "What is the total amount for February?": "Total Amount: 70",
        }
        for question, expected in cases.items():
            with self.subTest(question=question):
                answer = analyze_workbook_question(question, 1, document_id=document_id)
                self.assertTrue(answer["grounded"])
                self.assertIn(expected, answer["answer"])

        with patch("app.services.rag_service.search_chunks", side_effect=AssertionError("must not vector search")), patch(
            "app.services.rag_service.generate_answer", side_effect=AssertionError("must not call LLM")
        ), patch("app.services.rag_service.log_audit_event"):
            answer = answer_question("What is the total amount for February?", 1, document_id=document_id)
        self.assertTrue(answer["grounded"])
        self.assertIn("Total Amount: 70", answer["answer"])
        self.assertEqual(
            {source["source_location"]["sheet_name"] for source in answer["sources"]},
            {"February"},
        )

    def test_structured_provenance_is_compact_and_explains_aggregation(self) -> None:
        """Answers expose calculation provenance, never serialized worksheet row contents."""
        result = self.upload_workbook(
            [("April", [["Region", "Amount"], ["North", 12], ["South", 18]], "visible")],
            "regional.xlsx",
        )
        answer = analyze_workbook_question("What is the total amount for North?", 1, document_id=int(result["document_id"]))

        provenance = answer["provenance"]
        self.assertEqual(provenance["document_id"], int(result["document_id"]))
        self.assertEqual(provenance["workbook_filename"], "regional.xlsx")
        self.assertEqual(provenance["sheets"], [{"sheet_name": "April", "row_ranges": [{"row_start": 2, "row_end": 2}]}])
        self.assertEqual(provenance["columns_used"], ["Amount", "Region"])
        self.assertEqual(provenance["filters_applied"], {"Region": ["north"]})
        self.assertEqual(provenance["aggregation"], "total")
        self.assertEqual(provenance["contributing_row_count"], 1)
        self.assertNotIn("rows", provenance)
        self.assertNotIn("North", str(provenance))
        self.assertEqual(answer["_context"]["result_plan"], provenance)

    def test_aggregation_selects_matching_workbook_and_respects_acl_lifecycle(self) -> None:
        """Structured aggregation honors document selection, ACLs, deletion, and current status."""
        self.upload_workbook(
            [("Catalog", [["Code", "Label"], ["A-1", "Alpha"]], "visible")],
            "catalog.xlsx",
        )
        target = self.upload_workbook(
            [("Ledger", [["Period", "Amount"], ["March", 25], ["April", 35]], "visible")],
            "ledger.xlsx",
        )
        document_id = int(target["document_id"])

        selected = analyze_workbook_question("What is the total amount for March?", 1)
        self.assertTrue(selected["grounded"])
        self.assertIn("25", selected["answer"])
        self.assertEqual(selected["sources"][0]["document_id"], document_id)

        self.assertFalse(analyze_workbook_question("What is the total amount?", 2, document_id=document_id)["grounded"])
        with database.get_connection() as connection:
            connection.execute("UPDATE documents SET deleted_at = CURRENT_TIMESTAMP WHERE id = ?", (document_id,))
        self.assertFalse(analyze_workbook_question("What is the total amount?", 1, document_id=document_id)["grounded"])

        replacement = self.upload_workbook(
            [("Ledger", [["Period", "Amount"], ["March", 99]], "visible")],
            "replacement.xlsx",
        )
        replacement_id = int(replacement["document_id"])
        with database.get_connection() as connection:
            connection.execute(
                "UPDATE document_versions SET status = 'failed' WHERE id = (SELECT current_version_id FROM documents WHERE id = ?)",
                (replacement_id,),
            )
        self.assertFalse(analyze_workbook_question("What is the total amount?", 1, document_id=replacement_id)["grounded"])


if __name__ == "__main__":
    unittest.main()
