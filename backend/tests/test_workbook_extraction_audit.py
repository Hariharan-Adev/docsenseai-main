"""Regression coverage for workbook extraction invariants across sheets."""

from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook

from app.services.source_extraction import extract_source_chunks
from app.services.workbooks import extract_workbook, workbook_chunks


class WorkbookExtractionAuditTests(TestCase):
    """Verify workbook and source extraction preserve stable per-sheet evidence."""

    def test_duplicate_headers_receive_stable_source_labels(self) -> None:
        """Both Excel readers use the same non-ambiguous header convention."""
        from app.services.source_extraction import _deduplicated_headers

        self.assertEqual(
            _deduplicated_headers([(1, "Amount"), (2, "Amount"), (3, " ")]),
            {1: "Amount", 2: "Amount (2)"},
        )

    def test_multisheet_workbook_preserves_structured_and_source_provenance(self) -> None:
        """Retain names, rows, types, merged data, formulas, and distinct headers."""
        with TemporaryDirectory() as directory:
            path = Path(directory) / "audit.xlsx"
            workbook = Workbook()
            north = workbook.active
            north.title = "North"
            north.append(["Code", "Amount", "Amount", "When"])
            north.append(["N-1", 10, "=B2*2", date(2026, 1, 2)])
            north.append(["", "", "", ""])
            south = workbook.create_sheet("South")
            south.append(["Category", "Units"])
            south.merge_cells("A2:A3")
            south["A2"] = "Hardware"
            south["B2"] = 4
            south["B3"] = 6
            workbook.save(path)
            workbook.close()

            structured = extract_workbook(path)
            chunks = extract_source_chunks(path)

        self.assertEqual([sheet.name for sheet in structured.sheets], ["North", "South"])
        north_sheet, south_sheet = structured.sheets
        self.assertEqual(north_sheet.headers, ["Code", "Amount", "Amount (2)", "When"])
        self.assertEqual(north_sheet.rows[0].row_number, 2)
        self.assertEqual(north_sheet.rows[0].values["Amount"], 10)
        self.assertEqual(north_sheet.rows[0].values["When"], "2026-01-02 00:00:00")
        self.assertEqual(len(north_sheet.rows), 1)
        self.assertEqual([row.values["Category"] for row in south_sheet.rows], ["Hardware", "Hardware"])
        self.assertEqual([row.row_number for row in south_sheet.rows], [2, 3])

        north_formula = next(chunk for chunk in chunks if chunk.location["cell_range"] == "A2:D2")
        self.assertIn("Amount: 10", north_formula.text)
        self.assertIn("Amount (2): =B2*2", north_formula.text)
        self.assertEqual(north_formula.location["formulas"], {"C2": "=B2*2"})
        self.assertEqual(north_formula.location["sheet_name"], "North")
        self.assertEqual(north_formula.location["row_start"], 2)
        self.assertTrue(all(chunk.location["sheet_name"] in {"North", "South"} for chunk in chunks))
        self.assertFalse(any(chunk.location["row_start"] == 3 and chunk.location["sheet_name"] == "North" for chunk in chunks))
        self.assertTrue(any("Sheet: South" in text for text, _, _ in workbook_chunks(structured, path.name)))
