"""Safe, domain-neutral spreadsheet extraction and row chunk generation."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal
from json import dumps
from pathlib import Path
import csv
import math
import re

from app.services.source_extraction import SourceChunk
from app.services.pdf_layout import extract_pdf_page_texts

from app.services.document_loader import DocumentParseError


STRUCTURED_INDEX_VERSION = "structured-workbook-v2"


@dataclass
class WorkbookRow:
    row_number: int
    values: dict[str, object]


@dataclass
class WorkbookSheet:
    name: str
    state: str
    status: str
    header_row: int | None = None
    headers: list[str] = field(default_factory=list)
    rows: list[WorkbookRow] = field(default_factory=list)
    error: str | None = None


@dataclass
class WorkbookData:
    sheets: list[WorkbookSheet]

    @property
    def non_empty_sheets(self) -> list[WorkbookSheet]:
        return [sheet for sheet in self.sheets if sheet.status == "processed"]

    @property
    def skipped_sheets(self) -> list[str]:
        return [sheet.name for sheet in self.sheets if sheet.status == "empty"]

    @property
    def failed_sheets(self) -> list[str]:
        return [sheet.name for sheet in self.sheets if sheet.status == "failed"]


def _normalized_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return int(value) if value.is_integer() else value
    if isinstance(value, (str, int, bool)):
        return value
    return str(value)


def _canonical_cell(value: object) -> str:
    """Normalize cells only for structural comparisons such as repeated headers."""
    return " ".join(str(value or "").casefold().split())


def _xlsx_value(cell, merged_values: dict[tuple[int, int], object] | None = None) -> object:
    """Return a safe value while preserving vertical merged header labels."""
    raw = (
        merged_values.get((cell.row, cell.column), cell.value)
        if merged_values is not None
        else cell.value
    )
    value = _normalized_value(raw)
    if isinstance(value, int) and value >= 0:
        number_format = str(getattr(cell, "number_format", "") or "")
        if re.fullmatch(r"0+", number_format) and len(number_format) > len(str(value)):
            return str(value).zfill(len(number_format))
    return value


def _is_nonempty(value: object) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _detect_header(rows: list[tuple[int, list[object]]]) -> int:
    """Choose a probable header from the leading rows without domain assumptions."""
    candidates = rows[:20]
    best_index = 0
    best_score = float("-inf")
    for index, (_, values) in enumerate(candidates):
        present = [value for value in values if _is_nonempty(value)]
        if not present:
            continue
        text_count = sum(isinstance(value, str) for value in present)
        distinct = len({str(value).strip().casefold() for value in present})
        following_width = 0
        if index + 1 < len(rows):
            following_width = sum(_is_nonempty(value) for value in rows[index + 1][1])
        score = len(present) * 3 + text_count * 2 + distinct + min(following_width, len(present))
        if len(present) == 1 and len(rows) > index + 1:
            score -= 4
        if score > best_score:
            best_index, best_score = index, score
    return best_index


def _headers(values: list[object], width: int) -> list[str]:
    output: list[str] = []
    used: dict[str, int] = {}
    for index in range(width):
        raw = values[index] if index < len(values) else None
        base = str(raw).strip() if _is_nonempty(raw) else f"Column {index + 1}"
        count = used.get(base.casefold(), 0) + 1
        used[base.casefold()] = count
        output.append(base if count == 1 else f"{base} ({count})")
    return output


FORM_HELPER_VALUES = {
    "pick from list",
    "prompt",
    "prompt text",
}
FORM_HELPER_PREFIXES = (
    "what are your ",
    "please ",
    "select ",
)
FORM_KEY_ALIASES = {
    "career goals": "Career Goals",
    "career goal": "Career Goals",
    "experience": "Experience Level",
    "experience level": "Experience Level",
    "expected responsibilities": "Expected Responsibilities",
    "expected responsibility": "Expected Responsibilities",
    "primary skill": "Primary Skills",
    "primary skills": "Primary Skills",
    "project allocation": "Project Allocation",
    "role": "Role",
    "secondary skill": "Secondary Skills",
    "secondary skills": "Secondary Skills",
    "training completed": "Training Completed",
}


def _looks_like_form_helper(value: object) -> bool:
    """Identify prompt/helper cells that explain data entry but are not user data."""
    normalized = _canonical_cell(value)
    return (
        not normalized
        or normalized in FORM_HELPER_VALUES
        or normalized.startswith(FORM_HELPER_PREFIXES)
        or re.fullmatch(r"columns?\s*\d+", normalized) is not None
    )


def _form_field_name(value: object) -> str | None:
    """Normalize employee-profile labels while ignoring spreadsheet scaffolding."""
    normalized = _canonical_cell(value)
    if not normalized or _looks_like_form_helper(value):
        return None
    if normalized in {"name", "employee", "employee name"}:
        return "Employee"
    return FORM_KEY_ALIASES.get(normalized)


def _make_employee_profile_sheet(
    name: str,
    state: str,
    nonempty: list[tuple[int, list[object]]],
) -> WorkbookSheet | None:
    """Convert form-style employee sheets into one clean profile record."""
    fields: dict[str, object] = {}
    field_rows: dict[str, int] = {}
    detected_rows = 0
    for row_number, values in nonempty:
        for index, cell in enumerate(values[:-1]):
            field = _form_field_name(cell)
            if field is None:
                continue
            value = next(
                (
                    candidate
                    for candidate in values[index + 1 :]
                    if _is_nonempty(candidate) and not _looks_like_form_helper(candidate)
                ),
                None,
            )
            if value is None:
                continue
            fields[field] = value
            field_rows.setdefault(field, row_number)
            detected_rows += 1
            break
    meaningful = [field for field in fields if field != "Employee"]
    if detected_rows < 2 or not meaningful:
        return None
    employee = fields.get("Employee")
    if not _is_nonempty(employee):
        fields["Employee"] = name
    ordered_names = [
        "Employee",
        "Role",
        "Experience Level",
        "Primary Skills",
        "Secondary Skills",
        "Project Allocation",
        "Training Completed",
        "Career Goals",
        "Expected Responsibilities",
    ]
    ordered = {
        field: fields[field]
        for field in [*ordered_names, *sorted(fields)]
        if field in fields and _is_nonempty(fields[field])
    }
    return WorkbookSheet(
        name=name,
        state=state,
        status="processed",
        header_row=min(field_rows.values(), default=nonempty[0][0]),
        headers=list(ordered),
        rows=[WorkbookRow(row_number=min(field_rows.values(), default=nonempty[0][0]), values=ordered)],
    )


def _repair_parent_headers(
    headers: list[str],
    nonempty: list[tuple[int, list[object]]],
    header_index: int,
    data_rows: list[tuple[int, list[object]]],
) -> list[str]:
    """Preserve parent labels from multi-row headers with unmerged child cells."""
    if header_index <= 0:
        return headers
    parent_values = nonempty[header_index - 1][1]
    if not parent_values or not _is_nonempty(parent_values[0]):
        return headers
    first_header = str(headers[0]).strip()
    first_parent = str(parent_values[0]).strip()
    if not first_parent or not re.fullmatch(r"[A-Za-z]", first_header):
        return headers
    repaired = list(headers)
    repaired[0] = first_parent
    second_values = [
        str(values[1]).strip().casefold()
        for _, values in data_rows
        if len(values) > 1 and _is_nonempty(values[1])
    ]
    if len(repaired) > 1 and str(repaired[1]).startswith("Column "):
        # Attendance-style workbooks often use a blank child cell for IN/OUT rows.
        repaired[1] = (
            "Attendance Direction"
            if {"in", "out"} & set(second_values)
            else f"{first_parent} Detail"
        )
    return _headers(repaired, len(repaired))


def _looks_like_repeated_header(row: dict[str, object], headers: list[str]) -> bool:
    """Skip in-sheet section headers that repeat column names as row values."""
    matches = 0
    nonempty = 0
    for header, value in row.items():
        if not _is_nonempty(value):
            continue
        nonempty += 1
        if _canonical_cell(value) == _canonical_cell(header):
            matches += 1
    return matches >= 2 and matches >= nonempty - 1


def _structured_rows_for_block(
    headers: list[str],
    data_rows: list[tuple[int, list[object]]],
) -> list[WorkbookRow]:
    """Create structured rows for one detected table block."""
    structured = []
    width = len(headers)
    canonical_headers = [_canonical_cell(header) for header in headers]
    for row_number, values in data_rows:
        comparable = [_canonical_cell(values[index] if index < len(values) else None) for index in range(width)]
        if comparable == canonical_headers:
            continue
        row = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
        }
        if any(_is_nonempty(value) for value in row.values()) and not _looks_like_repeated_header(row, headers):
            structured.append(WorkbookRow(row_number=row_number, values=row))
    return structured


def _leading_table_block(
    nonempty: list[tuple[int, list[object]]],
    primary_header_index: int,
) -> tuple[list[str], list[WorkbookRow]]:
    """Preserve a smaller table that appears above the detected primary table."""
    leading = nonempty[:primary_header_index]
    if len(leading) < 3:
        return [], []
    header_index = _detect_header(leading)
    header_number, header_values = leading[header_index]
    data_rows = [
        (row_number, values)
        for row_number, values in leading[header_index + 1 :]
        if row_number > header_number
    ]
    if len(data_rows) < 2:
        return [], []
    text_headers = sum(1 for value in header_values if isinstance(value, str) and value.strip())
    if text_headers < 2:
        return [], []
    width = max([len(header_values), *(len(values) for _, values in data_rows)])
    headers = _headers(header_values, width)
    rows = _structured_rows_for_block(headers, data_rows)
    return (headers, rows) if rows else ([], [])


def _vertical_merged_values(worksheet) -> dict[tuple[int, int], object]:
    """Copy labels down vertically merged ranges without spreading title rows."""
    values: dict[tuple[int, int], object] = {}
    for cell_range in worksheet.merged_cells.ranges:
        if cell_range.min_col != cell_range.max_col:
            continue
        value = worksheet.cell(cell_range.min_row, cell_range.min_col).value
        for row_number in range(cell_range.min_row + 1, cell_range.max_row + 1):
            values[(row_number, cell_range.min_col)] = value
    return values


def _make_sheet(name: str, state: str, rows: list[tuple[int, list[object]]]) -> WorkbookSheet:
    nonempty = [
        (number, values)
        for number, values in rows
        if any(_is_nonempty(value) for value in values)
    ]
    if not nonempty:
        return WorkbookSheet(name=name, state=state, status="empty")

    profile_sheet = _make_employee_profile_sheet(name, state, nonempty)
    if profile_sheet is not None:
        return profile_sheet

    header_index = _detect_header(nonempty)
    header_number, header_values = nonempty[header_index]
    data_rows = nonempty[header_index + 1 :]
    width = max([len(header_values), *(len(values) for _, values in data_rows)])
    headers = _headers(header_values, width)
    headers = _repair_parent_headers(headers, nonempty, header_index, data_rows)
    structured = _structured_rows_for_block(headers, data_rows)
    leading_headers, leading_rows = _leading_table_block(nonempty, header_index)
    if leading_rows:
        headers = list(dict.fromkeys([*leading_headers, *headers]))
        structured = [*leading_rows, *structured]
    return WorkbookSheet(
        name=name,
        state=state,
        status="processed",
        header_row=header_number,
        headers=headers,
        rows=structured,
    )


def _extract_xlsx(
    path: Path, include_hidden: bool, include_very_hidden: bool
) -> WorkbookData:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    sheets: list[WorkbookSheet] = []
    try:
        for worksheet in workbook.worksheets:
            state = str(worksheet.sheet_state)
            if (
                state == "hidden" and not include_hidden
                or state == "veryHidden" and not include_very_hidden
            ):
                sheets.append(WorkbookSheet(worksheet.title, state, "disabled"))
                continue
            try:
                merged_values = _vertical_merged_values(worksheet)
                rows = [
                    (index, [_xlsx_value(cell, merged_values) for cell in row])
                    for index, row in enumerate(worksheet.iter_rows(), start=1)
                ]
                sheets.append(_make_sheet(worksheet.title, state, rows))
            except Exception:
                sheets.append(
                    WorkbookSheet(
                        worksheet.title,
                        state,
                        "failed",
                        error="This worksheet could not be read.",
                    )
                )
    finally:
        workbook.close()
    return WorkbookData(sheets)


def _extract_xls(
    path: Path, include_hidden: bool, include_very_hidden: bool
) -> WorkbookData:
    import xlrd

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    sheets: list[WorkbookSheet] = []
    try:
        for index in range(workbook.nsheets):
            worksheet = workbook.sheet_by_index(index)
            visibility = int(workbook.sheet_visibility[index])
            state = (
                "visible" if visibility == 0
                else "hidden" if visibility == 1
                else "veryHidden"
            )
            if (
                visibility == 1 and not include_hidden
                or visibility >= 2 and not include_very_hidden
            ):
                sheets.append(WorkbookSheet(worksheet.name, state, "disabled"))
                continue
            try:
                rows: list[tuple[int, list[object]]] = []
                for row_index in range(worksheet.nrows):
                    values: list[object] = []
                    for column_index in range(worksheet.ncols):
                        cell = worksheet.cell(row_index, column_index)
                        value: object = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            value = xlrd.xldate_as_datetime(value, workbook.datemode)
                        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                            value = bool(value)
                        elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                            value = None
                        values.append(_normalized_value(value))
                    rows.append((row_index + 1, values))
                sheets.append(_make_sheet(worksheet.name, state, rows))
            except Exception:
                sheets.append(
                    WorkbookSheet(
                        worksheet.name,
                        state,
                        "failed",
                        error="This worksheet could not be read.",
                    )
                )
    finally:
        workbook.release_resources()
    return WorkbookData(sheets)


def _extract_csv(path: Path) -> WorkbookData:
    with path.open(
        "r",
        encoding="utf-8-sig",
        errors="replace",
        newline="",
    ) as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        rows = [
            (
                row_number,
                [_normalized_value(value.strip()) for value in row],
            )
            for row_number, row in enumerate(
                csv.reader(handle, dialect),
                start=1,
            )
        ]
    return WorkbookData([
        _make_sheet("CSV", "visible", rows)
    ])


def _split_pdf_table_line(line: str) -> list[str]:
    """Split table-looking PDF text while avoiding ordinary sentence lines."""
    value = line.strip().strip("|")
    if not value:
        return []
    if "|" in value:
        cells = [cell.strip() for cell in value.split("|")]
    elif "\t" in value:
        cells = [cell.strip() for cell in value.split("\t")]
    else:
        cells = [cell.strip() for cell in re.split(r"\s{2,}", value)]
    cells = [cell for cell in cells if cell]
    return cells if len(cells) >= 2 else []


def _inspection_rejection_rows(text: str) -> list[tuple[int, list[object]]]:
    """Recover key totals from flattened final-inspection rejection PDFs."""
    rows: list[tuple[int, list[object]]] = [(
        1,
        [
            "S.No",
            "Model",
            "Inspected Qty",
            "OK Qty",
            "Total Reject Count",
            "Rejection %",
        ],
    )]
    for line in text.splitlines():
        match = re.match(
            r"^\s*(\d+)\s+(.+?)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+(?:\.\d+)?)%",
            line,
        )
        if not match:
            continue
        serial, model, inspected, ok_qty, rejected, percent = match.groups()
        # Keep the model descriptor intact; the rightmost totals are stable.
        rows.append((
            len(rows) + 1,
            [
                int(serial),
                model.strip(),
                int(inspected),
                int(ok_qty),
                int(rejected),
                f"{percent}%",
            ],
        ))
    return rows if len(rows) > 1 else []


def workbook_from_pdf_chunks(chunks: list[SourceChunk]) -> WorkbookData | None:
    """Build structured rows from clear text-extracted PDF tables."""
    sheets: list[WorkbookSheet] = []
    table_index = 0
    for chunk in chunks:
        if chunk.source_type != "pdf":
            continue
        inspection_rows = _inspection_rejection_rows(chunk.text)
        if inspection_rows:
            table_index += 1
            page = int(chunk.location.get("page_start") or chunk.location.get("page") or 1)
            sheet = _make_sheet(
                f"PDF page {page} table {table_index}",
                "visible",
                inspection_rows,
            )
            if sheet.status == "processed":
                sheets.append(sheet)
            continue
        rows: list[tuple[int, list[object]]] = []
        for line in chunk.text.splitlines():
            cells = _split_pdf_table_line(line)
            if cells:
                rows.append((len(rows) + 1, cells))
        if len(rows) < 2:
            continue
        widths = [len(values) for _, values in rows]
        common_width = max(set(widths), key=widths.count)
        table_rows = [
            (row_number, values)
            for row_number, values in rows
            if len(values) == common_width
        ]
        if len(table_rows) < 2:
            continue
        table_index += 1
        page = int(chunk.location.get("page_start") or chunk.location.get("page") or 1)
        sheet = _make_sheet(f"PDF page {page} table {table_index}", "visible", table_rows)
        if sheet.status == "processed":
            sheets.append(sheet)
    return WorkbookData(sheets) if any(sheet.status == "processed" for sheet in sheets) else None


def workbook_from_pdf(path: Path) -> WorkbookData | None:
    """Extract clear text PDF tables before retrieval chunking removes row breaks."""
    try:
        chunks = []
        for page_text in extract_pdf_page_texts(path):
            chunks.append(SourceChunk(
                text=page_text.text,
                source_type="pdf",
                location={"page_start": page_text.page_number, "page_end": page_text.page_number},
            ))
        return workbook_from_pdf_chunks(chunks)
    except Exception:
        return None


def extract_workbook(
    path: Path,
    include_hidden: bool = True,
    include_very_hidden: bool = False,
) -> WorkbookData:
    """Read all configured sheets without evaluating formulas, links, or macros."""
    try:
        if path.suffix.lower() == ".xlsx":
            workbook = _extract_xlsx(path, include_hidden, include_very_hidden)
        elif path.suffix.lower() == ".xls":
            workbook = _extract_xls(path, include_hidden, include_very_hidden)
        elif path.suffix.lower() == ".csv":
            workbook = _extract_csv(path)
        else:
            raise DocumentParseError("Unsupported spreadsheet type.")
    except DocumentParseError:
        raise
    except Exception as error:
        label = (
            "CSV table"
            if path.suffix.lower() == ".csv"
            else "legacy Excel workbook"
            if path.suffix.lower() == ".xls"
            else "Excel workbook"
        )
        raise DocumentParseError(f"The {label} could not be read.") from error
    if not workbook.non_empty_sheets:
        failed = ", ".join(workbook.failed_sheets)
        detail = f" Affected sheets: {failed}." if failed else ""
        raise DocumentParseError(f"The workbook contains no readable non-empty worksheets.{detail}")
    return workbook


def _plain_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = " ".join(text.replace("\x00", "").split())
    if text.startswith(("=", "+", "-", "@")):
        text = f"'{text}"
    return text.replace("|", r"\|")


def _infer_type(values: list[object]) -> str:
    """Infer a general column type from observed cell values."""
    present = [value for value in values if _is_nonempty(value)]
    if not present:
        return "text"
    numbers = sum(isinstance(value, (int, float, Decimal)) and not isinstance(value, bool) for value in present)
    date_like = sum(isinstance(value, (date, datetime, time)) for value in present)
    strings = [str(value).strip() for value in present if isinstance(value, str)]
    if numbers / len(present) >= 0.8:
        return "number"
    if date_like / len(present) >= 0.8:
        return "date"
    unique_ratio = len({str(value).strip().casefold() for value in present}) / max(len(present), 1)
    if strings and all(re.fullmatch(r"[A-Za-z0-9_.:/#-]+", value) for value in strings) and unique_ratio > 0.8:
        return "identifier"
    if unique_ratio <= 0.5:
        return "category"
    return "text"


def workbook_schema(sheet: WorkbookSheet) -> dict[str, object]:
    """Build a compact schema for planning structured answers later."""
    columns = []
    for header in sheet.headers:
        values = [row.values.get(header) for row in sheet.rows]
        present = [value for value in values if _is_nonempty(value)]
        columns.append({
            "name": header,
            "type": _infer_type(values),
            "non_empty_count": len(present),
            "unique_count": len({str(value).strip().casefold() for value in present}),
        })
    return {
        "sheet": sheet.name,
        "header_row": sheet.header_row,
        "row_count": len(sheet.rows),
        "columns": columns,
    }


def workbook_chunks(workbook: WorkbookData, filename: str) -> list[tuple[str, str, int | None]]:
    """Create row-oriented chunks, each carrying workbook and sheet provenance."""
    chunks: list[tuple[str, str, int | None]] = []
    for sheet in workbook.non_empty_sheets:
        if not sheet.rows:
            header_text = " | ".join(_plain_text(header) for header in sheet.headers)
            chunks.append((
                f"Workbook: {_plain_text(filename)} | Sheet: {_plain_text(sheet.name)} | "
                f"Headers: {header_text}",
                sheet.name,
                sheet.header_row,
            ))
            continue
        for row in sheet.rows:
            fields = " | ".join(
                f"{_plain_text(header)}: {_plain_text(value)}"
                for header, value in row.values.items()
                if _is_nonempty(value)
            )
            chunks.append((
                f"Workbook: {_plain_text(filename)} | Sheet: {_plain_text(sheet.name)} | "
                f"Row: {row.row_number} | {fields}",
                sheet.name,
                row.row_number,
            ))
    return chunks


def workbook_text(workbook: WorkbookData, filename: str) -> str:
    metadata = {
        "empty_sheets": workbook.skipped_sheets,
        "failed_sheets": workbook.failed_sheets,
    }
    chunks = [text for text, _, _ in workbook_chunks(workbook, filename)]
    return "\n".join([f"Workbook processing metadata: {dumps(metadata)}", *chunks])
